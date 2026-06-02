#!/usr/bin/env python3
"""
(주)비엔코 영업이익 시뮬레이터
버전: 1.0.0
작성일: 2026-05-11

사용법:
  pip install -r requirements.txt
  streamlit run app.py

임직원 공유:
  같은 네트워크: http://[서버IP]:8501
  외부 접속: Streamlit Community Cloud 배포 권장
"""

import streamlit as st
import pandas as pd
import altair as alt
from datetime import datetime, date
from typing import Optional, List
import calendar
import os
import re
import json
from pathlib import Path

# app.py 파일 위치 기준으로 절대경로 설정 (작업 디렉토리와 무관하게 동작)
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
CREDENTIALS_PATH = os.path.join(SCRIPT_DIR, 'credentials.json')

# ─── 페이지 설정 ───────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="비엔코 영업이익 시뮬레이터",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ─── CSS 스타일 ────────────────────────────────────────────────────────────────
st.markdown("""
<style>
    [data-testid="stSidebar"] { background-color: #f8f9fa; }
    [data-testid="stSidebar"] h1, [data-testid="stSidebar"] h2,
    [data-testid="stSidebar"] h3 { font-size: 0.95rem !important; }
    .scenario-box {
        border-radius: 10px; padding: 16px; margin: 4px 0;
        border: 1px solid #dee2e6;
    }
    .scenario-box.featured { border: 2px solid #0d6efd; }
    .profit-pos { color: #198754; font-weight: 700; }
    .profit-neg { color: #dc3545; font-weight: 700; }
    .avg-badge {
        background: #e9f5ff; color: #0a58ca; font-size: 0.75rem;
        padding: 2px 6px; border-radius: 4px; margin-left: 4px;
    }
    .section-divider { margin: 1.2rem 0; border-top: 1px solid #dee2e6; }
    div[data-testid="metric-container"] {
        background: #f8f9fa; border-radius: 8px; padding: 10px;
    }
</style>
""", unsafe_allow_html=True)

# ─── 구글 시트 설정 ────────────────────────────────────────────────────────────
SHEET_IDS = {
    'daily_sales': '1N3x5jBQ9ZY9M8hsyn0rnnvgwG6K5OydU6i4LtixIL4o',
    'historical_pl': '15D6YUXeZ0-ycHaJt7eEi9uGydbIh7T73Q_vlAwF8Jlg',
    'estimated':    '1VhkWW10TbSvsypkyOblrvAl0lxuHwAl4sFAzKmwn9Jo',
}
SCOPES = [
    'https://spreadsheets.google.com/feeds',
    'https://www.googleapis.com/auth/drive',
]

# ─── 구글 시트 연결 ────────────────────────────────────────────────────────────
_GC_ERROR = ""  # 전역 오류 메시지

@st.cache_resource(show_spinner=False)
def get_gspread_client():
    """서비스 계정으로 gspread 클라이언트 초기화"""
    global _GC_ERROR
    try:
        import gspread
        from google.oauth2.service_account import Credentials

        if os.path.exists(CREDENTIALS_PATH):
            creds = Credentials.from_service_account_file(CREDENTIALS_PATH, scopes=SCOPES)
        elif 'google_credentials' in st.secrets:
            creds = Credentials.from_service_account_info(
                dict(st.secrets['google_credentials']), scopes=SCOPES
            )
        else:
            _GC_ERROR = f"파일 없음: {CREDENTIALS_PATH}"
            return None
        client = gspread.authorize(creds)
        _GC_ERROR = ""
        return client
    except Exception as e:
        _GC_ERROR = str(e)
        return None


# ─── 데이터 로딩 ───────────────────────────────────────────────────────────────
@st.cache_data(ttl=300, show_spinner=False)
def load_sheet_data(sheet_id: str, tab_keyword: Optional[str] = None, by_gid: Optional[int] = None):
    """지정한 시트 탭 데이터를 2D 리스트로 반환

    탐색 우선순위:
      1. by_gid 지정 시 해당 gid 탭
      2. tab_keyword 지정 시 이름에 키워드 포함된 탭
      3. 모두 없으면 첫 번째 탭
    """
    gc = get_gspread_client()
    if gc is None:
        return None, "credentials.json 없음 — 아래 설정 가이드 참조"
    try:
        sh  = gc.open_by_key(sheet_id)
        all_ws = sh.worksheets()
        ws  = None
        if by_gid is not None:
            ws = next((w for w in all_ws if w.id == by_gid), None)
        if ws is None and tab_keyword:
            # 정확히 일치하는 탭 우선, 없으면 포함 탭
            ws = next((w for w in all_ws if w.title == tab_keyword), None)
            if ws is None:
                ws = next((w for w in all_ws if tab_keyword in w.title), None)
        if ws is None:
            ws = all_ws[0]
        return ws.get_all_values(), None
    except Exception as e:
        return None, str(e)


@st.cache_data(ttl=600, show_spinner=False)
def get_daily_sheet_tabs():
    """daily_sales 스프레드시트의 모든 탭 목록 반환 → [(title, gid), ...]"""
    gc = get_gspread_client()
    if gc is None:
        return []
    try:
        sh = gc.open_by_key(SHEET_IDS['daily_sales'])
        return [(w.title, w.id) for w in sh.worksheets()]
    except Exception:
        return []


# ─── 숫자 파싱 ─────────────────────────────────────────────────────────────────
def pn(s) -> float:
    """문자열에서 숫자 추출 (₩·쉼표·%·공백·음수기호 모두 처리)"""
    if s is None:
        return 0.0
    t = str(s).strip()
    if t in ['-', '', '\\-', '−', '₩ -', '₩-', '#REF!', '#N/A']:
        return 0.0
    try:
        # 숫자/소수점/마이너스 이외 모든 문자 제거
        cleaned = re.sub(r'[^\d.\-]', '', t.replace('−', '-').replace('\\-', '-'))
        return float(cleaned) if cleaned and cleaned not in ['-', '.'] else 0.0
    except Exception:
        return 0.0


def find_row(data, *keywords) -> Optional[list]:
    """키워드가 포함된 첫 번째 행 반환"""
    for row in data:
        row_text = ' '.join(str(c) for c in row)
        if all(kw in row_text for kw in keywords):
            return row
    return None


def monthly_values(row, skip=3, step=2) -> List[float]:
    """행에서 월별 금액 추출 (skip: 앞 컬럼 건너뜀, step: 금액/비율 교번)"""
    if row is None:
        return []
    return [pn(row[i]) for i in range(skip, len(row), step)]


def safe_avg(lst, last_n=3) -> float:
    """최근 n개 비0값 평균"""
    vals = [v for v in lst if v > 0][-last_n:]
    return sum(vals) / len(vals) if vals else 0.0


def safe_rate_avg(amounts, revenues, last_n=3) -> float:
    """최근 n개월 비율 평균 (%)"""
    pairs = [(a, r) for a, r in zip(amounts, revenues) if r > 0 and a >= 0][-last_n:]
    rates = [a / r * 100 for a, r in pairs]
    return sum(rates) / len(rates) if rates else 0.0


# ─── P&L 파싱 → 3개월 평균 계산 ──────────────────────────────────────────────
FALLBACK_AVGS = {
    'cogs_rate':       35.5,
    'adv_rate':        25.0,
    'commission_rate': 8.5,
    'delivery_rate':   8.0,
    'fee_rate':        2.4,
    'export_rate':     2.5,
    'fixed': {
        '인건비':      74_259_079,
        '복리후생비':   6_000_000,
        '여비교통비':   1_400_000,
        '접대비':         560_000,
        '지급임차료':     626_273,
        '보험료':         969_778,
        '통신비':         270_000,
        '감가상각비':     807_114,
        '용역수수료':     310_230,
    },
    'source': '기본값 (2025년 실적 기반)',
    'months': [],
}


def calculate_averages(pl_data) -> dict:
    """손익 시트 데이터로 직전 3개월 평균 계산"""
    if pl_data is None:
        return FALLBACK_AVGS.copy()
    try:
        rev_row  = find_row(pl_data, '매      출      액')
        cogs_row = find_row(pl_data, '매    출   원   가')
        adv_row  = find_row(pl_data, '광고선전비')
        comm_row = find_row(pl_data, '판매수수료')
        dlv_row  = find_row(pl_data, '운반비')
        fee_row  = find_row(pl_data, '지급수수료')
        exp_row  = find_row(pl_data, '수출제비용')
        sal_row  = find_row(pl_data, '직원급여')
        ret_row  = find_row(pl_data, '퇴직급여')
        wlf_row  = find_row(pl_data, '복리후생비')
        trv_row  = find_row(pl_data, '여비교통비')
        ent_row  = find_row(pl_data, '접대비')
        rnt_row  = find_row(pl_data, '지급임차료')
        ins_row  = find_row(pl_data, '보험료')
        tel_row  = find_row(pl_data, '통신비')
        dep_row  = find_row(pl_data, '감가상각비')
        svc_row  = find_row(pl_data, '용역수수료')

        revs  = monthly_values(rev_row)
        cogss = monthly_values(cogs_row)
        advs  = monthly_values(adv_row)
        comms = monthly_values(comm_row)
        dlvs  = monthly_values(dlv_row)
        fees  = monthly_values(fee_row)
        exps  = monthly_values(exp_row)

        # 월 레이블 구성 (행 헤더에서 추출)
        months = []
        if rev_row:
            header = pl_data[0] if pl_data else []
            for i in range(3, len(header), 2):
                if header[i]:
                    months.append(str(header[i]).strip())

        cogs_rate       = safe_rate_avg(cogss, revs)
        adv_rate        = safe_rate_avg(advs, revs)
        commission_rate = safe_rate_avg(comms, revs)
        delivery_rate   = safe_rate_avg(dlvs, revs)
        fee_rate        = safe_rate_avg(fees, revs)
        export_rate     = safe_rate_avg(exps, revs)

        def fa(row): return safe_avg(monthly_values(row))

        # 인건비 = 급여 + 퇴직
        salary_vals = [a + b for a, b in zip(monthly_values(sal_row), monthly_values(ret_row))]
        salary_avg  = safe_avg(salary_vals)

        return {
            'cogs_rate':       round(cogs_rate, 1) or FALLBACK_AVGS['cogs_rate'],
            'adv_rate':        round(adv_rate, 1) or FALLBACK_AVGS['adv_rate'],
            'commission_rate': round(commission_rate, 1) or FALLBACK_AVGS['commission_rate'],
            'delivery_rate':   round(delivery_rate, 1) or FALLBACK_AVGS['delivery_rate'],
            'fee_rate':        round(fee_rate, 1) or FALLBACK_AVGS['fee_rate'],
            'export_rate':     round(export_rate, 1) or FALLBACK_AVGS['export_rate'],
            'fixed': {
                '인건비':      round(salary_avg) or FALLBACK_AVGS['fixed']['인건비'],
                '복리후생비':  round(fa(wlf_row)) or FALLBACK_AVGS['fixed']['복리후생비'],
                '여비교통비':  round(fa(trv_row)) or FALLBACK_AVGS['fixed']['여비교통비'],
                '접대비':      round(fa(ent_row)) or FALLBACK_AVGS['fixed']['접대비'],
                '지급임차료':  round(fa(rnt_row)) or FALLBACK_AVGS['fixed']['지급임차료'],
                '보험료':      round(fa(ins_row)) or FALLBACK_AVGS['fixed']['보험료'],
                '통신비':      round(fa(tel_row)) or FALLBACK_AVGS['fixed']['통신비'],
                '감가상각비':  round(fa(dep_row)) or FALLBACK_AVGS['fixed']['감가상각비'],
                '용역수수료':  round(fa(svc_row)) or FALLBACK_AVGS['fixed']['용역수수료'],
            },
            'source': f"구글시트 직전 3개월 평균 ({', '.join(months[-3:]) if months else '최근'})",
            'months': months,
        }
    except Exception:
        return FALLBACK_AVGS.copy()


def parse_sheet_monthly_pl(pl_data) -> dict:
    """구글시트 역사 PL에서 월별 주요 손익 항목 추출
    반환: {'2026-01': {'매출액':..., '매출원가':..., ...}, ...}
    """
    if not pl_data:
        return {}
    try:
        header = pl_data[0] if pl_data else []
        months, month_cols = [], []
        for i in range(3, len(header), 2):
            v = header[i]
            if v and re.match(r'\d{4}-\d{2}', str(v).strip()):
                m = str(v).strip()
                if m not in months:
                    months.append(m)
                    month_cols.append(i)

        def gv(keyword):
            row = find_row(pl_data, keyword)
            if row is None:
                return {}
            return {m: pn(row[c]) if c < len(row) else 0.0
                    for m, c in zip(months, month_cols)}

        rev  = gv('매      출      액')
        cogs = gv('매    출   원   가')
        gp   = gv('매  출  총  이  익')
        sga  = gv('판  매  관  리  비')
        op   = gv('영업손')   # 영업손실/이익 공통 키워드
        adv  = gv('광고선전비')
        comm = gv('판매수수료')
        dlv  = gv('운반비')
        fee  = gv('지급수수료')
        exp  = gv('수출제비용')

        result = {}
        for m in months:
            r = rev.get(m, 0)
            if r <= 0:
                continue
            # 영업손익 부호: 손실행이면 음수
            op_val = op.get(m, 0)
            # gp - sga 로 재계산 (더 안정적)
            gp_val  = gp.get(m, 0) or (r - cogs.get(m, 0))
            sga_val = sga.get(m, 0)
            if gp_val and sga_val:
                op_val = gp_val - sga_val

            result[m] = {
                '매출액':    r,
                '매출원가':  cogs.get(m, 0),
                '매출총이익': gp_val,
                '판관비계':  sga_val,
                '영업손익':  op_val,
                '광고선전비': adv.get(m, 0),
                '판매수수료': comm.get(m, 0),
                '운반비':    dlv.get(m, 0),
                '지급수수료': fee.get(m, 0),
                '수출제비용': exp.get(m, 0),
            }
        return result
    except Exception:
        return {}


def simulate_month_pl(revenue: float, ss: dict) -> dict:
    """세션 파라미터로 월 손익 시뮬레이션"""
    cogs   = revenue * ss.get('p_cogs', 33) / 100
    gross  = revenue - cogs
    adv    = revenue * ss.get('p_adv', 20) / 100
    comm   = revenue * ss.get('p_comm', 10) / 100
    dlv    = revenue * ss.get('p_dlv', 8) / 100
    fee    = revenue * ss.get('p_fee', 2.4) / 100
    exp    = revenue * ss.get('p_exp', 2.5) / 100
    fixed  = sum(ss.get(k, 0) for k in
                 ['p_salary','p_welfare','p_travel','p_entmt',
                  'p_rent','p_ins','p_tel','p_dep','p_svc'])
    var_sg = adv + comm + dlv + fee + exp
    tot_sg = var_sg + fixed
    op     = gross - tot_sg
    return {
        '매출액':    revenue, '매출원가': cogs,
        '매출총이익': gross,  '변동판관비': var_sg,
        '고정판관비': fixed,  '판관비계': tot_sg,
        '영업손익':  op,
    }


# ─── 데일리 매출 파싱 ─────────────────────────────────────────────────────────
FALLBACK_SALES = {
    'total':         176_856_916,
    'b2c':            47_928_016,
    'b2b':           128_928_900,
    'daily':         [28029010, 4225908, 5745763, 35950127, 3742293,
                      76490304, 6039511, 8964252, 4172066, 3497680],
    'b2c_daily':     [6646760, 4225908, 5745763, 4868127, 3742293,
                      5190154, 5275011, 4564252, 4172066, 3497680],
    'b2b_daily':     [21382250, 0, 0, 31082000, 0,
                      71300150, 764500, 4400000, 0, 0],
    'days_with_data': 10,
    'source':        '데모 데이터 (2026-05-01 ~ 05-10)',
}


def parse_daily_sales(raw_data) -> dict:
    """당월 매출 시트 → 일별 실적 / 누계 파싱

    '합계매출' 셀의 컬럼 위치를 기준으로 동적으로 일별 컬럼을 탐색:
      합계매출_col + 1 : 월 누계
      합계매출_col + 2~6 : 주간 합계 (1W~5W)
      합계매출_col + 7~ : 일별 매출 (1일~31일)
    """
    if raw_data is None:
        return FALLBACK_SALES.copy()
    try:
        target_row = None
        anchor_idx = None   # 합계매출 행 인덱스 (B2B/B2C daily 추출용)
        합계_col   = None
        b2c_total  = 0
        b2b_total  = 0

        for ridx, row in enumerate(raw_data):
            for i, cell in enumerate(row):
                if '합계매출' in str(cell):
                    target_row = row
                    anchor_idx = ridx
                    합계_col   = i
                    break
            if target_row is not None:
                break

        if target_row is None or 합계_col is None:
            return FALLBACK_SALES.copy()

        # 월 누계: 합계_col + 1
        monthly_total = pn(target_row[합계_col + 1]) if len(target_row) > 합계_col + 1 else 0

        # 일별 데이터: 합계_col + 7 부터 최대 31개
        daily_start = 합계_col + 7
        daily_raw   = target_row[daily_start:daily_start + 31] if len(target_row) > daily_start else []

        # B2C / B2B 누계 추출 (col[3] 기준 탐색)
        b2c_row = None
        b2b_row = None
        for row in raw_data:
            if len(row) > 3:
                cs = str(row[3]).strip()
                if cs == 'B2C' and b2c_row is None:
                    b2c_row   = row
                    b2c_total = pn(row[합계_col + 1]) if len(row) > 합계_col + 1 else 0
                elif cs == 'B2B' and b2b_row is None:
                    b2b_row   = row
                    b2b_total = pn(row[합계_col + 1]) if len(row) > 합계_col + 1 else 0

        # 일별 금액 파싱
        daily_clean    = [pn(v) for v in daily_raw]
        # 입력된 마지막 날까지만 사용
        last_idx = -1
        for i, v in enumerate(daily_clean):
            if v > 0:
                last_idx = i

        # B2C / B2B 일별 데이터 추출
        def _extract_daily(row):
            if row is None:
                return []
            raw = row[daily_start:daily_start + 31] if len(row) > daily_start else []
            vals = [pn(v) for v in raw]
            last = -1
            for k, v in enumerate(vals):
                if v > 0:
                    last = k
            return vals[:last + 1] if last >= 0 else []

        b2c_daily = _extract_daily(b2c_row)
        b2b_daily = _extract_daily(b2b_row)

        if last_idx < 0:
            # 일별 컬럼에서 데이터를 못 찾으면 월 누계만 사용
            return {
                'total':          monthly_total if monthly_total > 0 else FALLBACK_SALES['total'],
                'b2c':            b2c_total,
                'b2b':            b2b_total,
                'daily':          [],
                'b2c_daily':      b2c_daily,
                'b2b_daily':      b2b_daily,
                'days_with_data': 0,
                'source':         'Google Sheets (일별 파싱 실패)',
            }

        daily_clean    = daily_clean[:last_idx + 1]
        days_with_data = len([v for v in daily_clean if v > 0])
        total          = sum(daily_clean)

        return {
            'total':          total if total > 0 else monthly_total,
            'b2c':            b2c_total,
            'b2b':            b2b_total,
            'daily':          daily_clean,
            'b2c_daily':      b2c_daily,
            'b2b_daily':      b2b_daily,
            'days_with_data': days_with_data,
            'source':         'Google Sheets 실시간',
        }
    except Exception:
        return FALLBACK_SALES.copy()


# ─── 매출 상세 파싱 ───────────────────────────────────────────────────────────
def parse_sales_detail(raw_data) -> Optional[dict]:
    """합계매출 행을 앵커로 B2C/B2B·브랜드·채널·국가별 매출 파싱

    앵커(합계매출행) 기준 상대 오프셋:
      -2 : B2C합계,  -1 : B2B합계
      +5~+9  : 누디크 채널별 (자사몰·스마트스토어·쿠팡·국군복지몰·임직원몰)
      +10    : 누디크 합계
      +11~+13: 스피큘엑스 채널별
      +15    : 스피큘엑스 합계
      +16~+17: 더콜린 채널별
      +19    : 더콜린 합계
      +20    : KO B2C합계
      +25    : JP B2C합계(한화)
      +29    : US B2C합계(원화)
      +31    : CN B2C합계(원화)
    """
    if not raw_data:
        return None
    try:
        # 앵커: col[3] == '합계매출'
        anchor = None
        for i, row in enumerate(raw_data):
            if len(row) > 3 and '합계매출' in str(row[3]):
                anchor = i
                break
        if anchor is None:
            return None

        def g(offset, col=4):
            idx = anchor + offset
            if 0 <= idx < len(raw_data) and col < len(raw_data[idx]):
                return pn(raw_data[idx][col])
            return 0.0

        # B2C / B2B
        b2c_total = g(-2)
        b2b_total = g(-1)

        # 브랜드별 (누디크·스피큘엑스·더콜린)
        brands = {
            '누디크':    g(+10),
            '스피큘엑스': g(+15),
            '더콜린':    g(+19),
        }

        # 채널별 (브랜드 내 합산)
        channels_by_brand = {
            '누디크': {
                '자사몰':       g(+5),
                '스마트스토어': g(+6),
                '쿠팡':         g(+7),
                '국군복지몰':   g(+8),
                '임직원몰':     g(+9),
            },
            '스피큘엑스': {
                '자사몰':       g(+11),
                '스마트스토어': g(+12),
                '쿠팡':         g(+13),
            },
            '더콜린': {
                '스마트스토어': g(+16),
                '쿠팡':         g(+17),
            },
        }

        # 채널 합산 (전체)
        channels: dict = {}
        for brand_ch in channels_by_brand.values():
            for ch, v in brand_ch.items():
                channels[ch] = channels.get(ch, 0) + v

        # 국가별 B2C (KO·JP·US·CN)
        countries = {
            'KO': g(+20),
            'JP': g(+25),
            'US': g(+29),
            'CN': g(+31),
        }

        # ── B2B 상세 동적 탐색 (anchor+32 이후 행 스캔) ──────────────────
        # 시트 구조: col[1]=국가, col[2]=브랜드, col[3]=벤더명, col[4]=매출합계
        # col[3] 벤더명이 없는 행은 소계/합계 행으로 간주 → skip
        b2b_details: List[dict] = []
        last_country = ''
        last_brand   = ''

        for off in range(32, 70):
            idx = anchor + off
            if idx >= len(raw_data):
                break
            row = raw_data[idx]
            if len(row) <= 4:
                continue
            val = pn(row[4])
            if val <= 0:
                continue

            c1 = str(row[1]).strip() if len(row) > 1 else ''
            c2 = str(row[2]).strip() if len(row) > 2 else ''
            c3 = str(row[3]).strip() if len(row) > 3 else ''

            # col[3] 벤더명이 없으면 소계/합계 행 → skip
            if not c3:
                if c1: last_country = c1
                if c2: last_brand   = c2
                continue

            # 국가·브랜드: 병합셀 해제 후에도 안전하게 이전 값 유지
            country = c1 if c1 else last_country
            brand   = c2 if c2 else last_brand

            if c1: last_country = c1
            if c2: last_brand   = c2

            b2b_details.append({
                'label':   c3,                          # 벤더명은 반드시 col[3]
                'brand':   brand if brand else '-',
                'country': country if country else 'KO',
                'value':   val,
            })

        # B2B 국가별/브랜드별 집계
        b2b_countries: dict = {}
        b2b_brands: dict    = {}
        for d in b2b_details:
            b2b_countries[d['country']] = b2b_countries.get(d['country'], 0) + d['value']
            if d['brand'] not in ('-', ''):
                b2b_brands[d['brand']] = b2b_brands.get(d['brand'], 0) + d['value']

        return {
            'b2c_total':         b2c_total,
            'b2b_total':         b2b_total,
            'brands':            brands,
            'channels':          channels,
            'channels_by_brand': channels_by_brand,
            'countries':         countries,
            'b2b_details':       b2b_details,
            'b2b_countries':     b2b_countries,
            'b2b_brands':        b2b_brands,
        }
    except Exception:
        return None



# ─── 재무제표 엑셀 파싱 ─────────────────────────────────────────────────────────
def parse_financial_excel(uploaded_file) -> Optional[dict]:
    """업로드된 재무제표 엑셀 → 손익(기간별) 시트 파싱
    
    반환 구조:
      months      : ['2026-01', '2026-02', ...]  (실제 데이터 있는 월)
      data        : {항목키: {월: 금액, '계': 합계}}
      sim_params  : 직전 3개월 평균 → 시뮬레이터 파라미터 dict
    """
    import io as _io
    try:
        import openpyxl
    except ImportError:
        return {'error': 'openpyxl 패키지 필요: pip install openpyxl'}
    try:
        # seek(0) 후 읽기 (Streamlit 업로드 파일 포인터 초기화)
        if hasattr(uploaded_file, 'seek'):
            uploaded_file.seek(0)
        raw = uploaded_file.read()

        # Strict OOXML → Transitional 변환 함수 (일부 Excel 저장 포맷 대응)
        import zipfile as _zf
        _STRICT_NS = [
            ("http://purl.oclc.org/ooxml/spreadsheetml/main",
             "http://schemas.openxmlformats.org/spreadsheetml/2006/main"),
            ("http://purl.oclc.org/ooxml/officeDocument/relationships",
             "http://schemas.openxmlformats.org/officeDocument/2006/relationships"),
            ("http://purl.oclc.org/ooxml/drawingml/2006/main",
             "http://schemas.openxmlformats.org/drawingml/2006/main"),
            ("http://purl.oclc.org/ooxml/drawingml/2006/spreadsheetDrawing",
             "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"),
            ("http://purl.oclc.org/ooxml/officeDocument/docPropsVTypes",
             "http://schemas.openxmlformats.org/officeDocument/2006/docPropsVTypes"),
        ]
        def _strict_to_transit(data):
            out = _io.BytesIO()
            with _zf.ZipFile(_io.BytesIO(data), 'r') as zin, \
                 _zf.ZipFile(out, 'w', _zf.ZIP_DEFLATED) as zout:
                for fname in zin.namelist():
                    b = zin.read(fname)
                    if fname.endswith('.xml') or fname.endswith('.rels'):
                        try:
                            t = b.decode('utf-8')
                            for old, new in _STRICT_NS:
                                t = t.replace(old, new)
                            b = t.encode('utf-8')
                        except Exception:
                            pass
                    zout.writestr(fname, b)
            return out.getvalue()

        # 워크북 로드: 먼저 그대로, 시트 없으면 Strict→Transit 변환 후 재시도
        wb = None
        for attempt_raw in [raw, _strict_to_transit(raw)]:
            for kwargs in [{'data_only': True}, {'data_only': True, 'keep_vba': True}]:
                try:
                    _wb = openpyxl.load_workbook(_io.BytesIO(attempt_raw), **kwargs)
                    if _wb.sheetnames:
                        wb = _wb
                        break
                except Exception:
                    pass
            if wb:
                break

        if wb is None or not wb.sheetnames:
            return {'error': f"엑셀 파일을 읽을 수 없습니다. 파일이 손상되었거나 지원되지 않는 형식입니다."}

        # '손익(기간별)' 시트 탐색
        # 전각괄호 ／ 반각괄호 ／ 공백 차이 무시하여 유연하게 검색
        def _norm_sheet(s):
            return re.sub(r'[\s　（）()]', '', str(s))

        ws = None
        for nm in wb.sheetnames:
            n = _norm_sheet(nm)
            if '기간별' in n or ('손익' in n and '기간' in n):
                ws = wb[nm]
                break
        if ws is None:
            # 마지막 fallback: 시트명 목록을 에러에 표시
            return {'error': f"손익(기간별) 시트 없음. 존재하는 시트: {wb.sheetnames}"}

        # ── 8행에서 월 컬럼 파악 (중복 제거) ──
        months, month_cols = [], []
        seen = set()
        for col in range(1, ws.max_column + 1):
            v = ws.cell(8, col).value
            if v and re.match(r"\d{4}-\d{2}", str(v).strip()):
                m = str(v).strip()
                if m not in seen:
                    months.append(m)
                    month_cols.append(col)
                    seen.add(m)
        if not months:
            return {'error': '월별 컬럼을 찾을 수 없습니다.'}

        def norm(s): return re.sub(r"\s+", "", str(s or ""))

        # ── 항목 키워드 매핑 (정규화된 레이블 검색) ──
        KEYWORDS = {
            "매출액":     ["매출액"],
            "매출원가":   ["매출원가"],
            "매출총이익": ["매출총이익"],
            "판관비계":   ["판매관리비"],
            "직원급여":   ["직원급여"],
            "상여금":     ["상여금"],
            "퇴직급여":   ["퇴직급여"],
            "복리후생비": ["복리후생비"],
            "여비교통비": ["여비교통비"],
            "접대비":     ["접대비"],
            "통신비":     ["통신비"],
            "감가상각비": ["감가상각비"],
            "지급임차료": ["지급임차료"],
            "보험료":     ["보험료"],
            "운반비":     ["운반비"],
            "소모품비":   ["소모품비"],
            "지급수수료": ["지급수수료"],
            "광고선전비": ["광고선전비"],
            "수출제비용": ["수출제비용"],
            "판매수수료": ["판매수수료"],
            "견본비":     ["견본비"],
            "용역수수료": ["용역수수료"],
            "영업손익":   ["영업손실", "영업이익"],
        }

        found = {k: None for k in KEYWORDS}
        for row in range(9, min(ws.max_row + 1, 120)):
            lv = ws.cell(row, 1).value
            if lv is None:
                continue
            label = norm(lv)
            for key, kws in KEYWORDS.items():
                if found[key] is not None:
                    continue
                if any(kw in label for kw in kws):
                    rd = {}
                    for m, mc in zip(months, month_cols):
                        v = ws.cell(row, mc).value
                        rd[m] = float(v) if isinstance(v, (int, float)) else 0.0
                    tot = ws.cell(row, 2).value
                    rd["계"] = float(tot) if isinstance(tot, (int, float)) else 0.0
                    # 영업손실은 음수로 저장
                    if key == "영업손익" and "손실" in label:
                        rd = {k: -abs(v) for k, v in rd.items()}
                    found[key] = rd
                    break

        # ── 실제 데이터 있는 월만 필터 ──
        active_months = [
            m for m in months
            if (found.get("매출액") or {}).get(m, 0) > 0
        ]
        if not active_months:
            active_months = months

        # ── 직전 3개월 평균 → 시뮬레이터 파라미터 ──
        last3 = active_months[-3:]

        def avg_rate(num_key, rev_key="매출액"):
            revs = [found.get(rev_key, {}).get(m, 0) for m in last3]
            nums = [found.get(num_key, {}).get(m, 0) for m in last3]
            pairs = [(n, r) for n, r in zip(nums, revs) if r > 0]
            if not pairs: return 0.0
            return sum(n / r * 100 for n, r in pairs) / len(pairs)

        def avg_fixed(key):
            vals = [found.get(key, {}).get(m, 0) for m in last3]
            vals = [v for v in vals if v > 0]
            return sum(vals) / len(vals) if vals else 0.0

        # 인건비 = 직원급여 + 상여금 + 퇴직급여 월 평균
        def avg_salary():
            total = []
            for m in last3:
                s = (found.get("직원급여", {}).get(m, 0)
                     + found.get("상여금", {}).get(m, 0)
                     + found.get("퇴직급여", {}).get(m, 0))
                total.append(s)
            vals = [v for v in total if v > 0]
            return sum(vals) / len(vals) if vals else 0.0

        sim_params = {
            "p_cogs":    round(avg_rate("매출원가"), 1),
            "p_adv":     round(avg_rate("광고선전비"), 1),
            "p_comm":    round(avg_rate("판매수수료"), 1),
            "p_dlv":     round(avg_rate("운반비"), 1),
            "p_fee":     round(avg_rate("지급수수료"), 1),
            "p_exp":     round(avg_rate("수출제비용"), 1),
            "p_salary":  round(avg_salary()),
            "p_welfare": round(avg_fixed("복리후생비")),
            "p_travel":  round(avg_fixed("여비교통비")),
            "p_entmt":   round(avg_fixed("접대비")),
            "p_rent":    round(avg_fixed("지급임차료")),
            "p_ins":     round(avg_fixed("보험료")),
            "p_tel":     round(avg_fixed("통신비")),
            "p_dep":     round(avg_fixed("감가상각비")),
            "p_svc":     round(avg_fixed("용역수수료")),
        }
        # 0인 항목은 현재 세션값 유지를 위해 제거
        sim_params = {k: v for k, v in sim_params.items() if v > 0}

        return {
            "months":       active_months,
            "all_months":   months,
            "data":         found,
            "sim_params":   sim_params,
            "last3":        last3,
            "filename":     getattr(uploaded_file, "name", "업로드 파일"),
        }
    except Exception as e:
        return {"error": str(e)}



# ─── 결산 히스토리 저장/로드 ─────────────────────────────────────────────────
_FIN_ITEMS = [
    "매출액","매출원가","매출총이익","판관비계",
    "직원급여","상여금","퇴직급여","복리후생비","여비교통비","접대비",
    "통신비","세금과공과금","감가상각비","지급임차료","보험료","운반비",
    "교육훈련비","소모품비","지급수수료","광고선전비","수출제비용",
    "판매수수료","견본비","용역수수료","영업손익",
]

def save_fin_month(month: str, filename: str, item_data: dict) -> None:
    """한 달치 손익 데이터를 JSON으로 저장 (덮어쓰기 허용)"""
    path = FIN_HISTORY_DIR / f"{month}.json"
    payload = {
        "month": month,
        "filename": filename,
        "uploaded_at": datetime.now().strftime("%Y-%m-%d %H:%M"),
        "data": {k: item_data.get(k, 0.0) for k in _FIN_ITEMS},
    }
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def load_fin_history() -> dict:
    """저장된 모든 월 데이터 로드 → {'YYYY-MM': {month,filename,uploaded_at,data}}"""
    history = {}
    for p in sorted(FIN_HISTORY_DIR.glob("*.json")):
        try:
            rec = json.loads(p.read_text(encoding="utf-8"))
            if "month" in rec and "data" in rec:
                history[rec["month"]] = rec
        except Exception:
            pass
    return history  # 월 기준 오름차순 정렬됨


# ─── 손익 계산 ─────────────────────────────────────────────────────────────────
def calc_pl(revenue: float, cogs_rate: float, adv_rate: float,
            comm_rate: float, delivery_rate: float, fee_rate: float,
            export_rate: float, fixed_total: float) -> dict:
    """매출 + 각 비율/금액 → 손익 계산"""
    cogs          = revenue * cogs_rate / 100
    gross         = revenue - cogs
    adv           = revenue * adv_rate / 100
    comm          = revenue * comm_rate / 100
    dlv           = revenue * delivery_rate / 100
    fee           = revenue * fee_rate / 100
    exp           = revenue * export_rate / 100
    var_sga       = adv + comm + dlv + fee + exp
    total_sga     = var_sga + fixed_total
    op_profit     = gross - total_sga
    safe_rev      = revenue if revenue > 0 else 1

    return {
        'revenue':          revenue,
        'cogs':             cogs,
        'cogs_rate':        cogs_rate,
        'gross':            gross,
        'gross_rate':       gross / safe_rev * 100,
        'adv':              adv,
        'comm':             comm,
        'dlv':              dlv,
        'fee':              fee,
        'exp':              exp,
        'var_sga':          var_sga,
        'fixed_sga':        fixed_total,
        'total_sga':        total_sga,
        'total_sga_rate':   total_sga / safe_rev * 100,
        'op_profit':        op_profit,
        'op_rate':          op_profit / safe_rev * 100,
    }


# ─── 포맷 유틸 ─────────────────────────────────────────────────────────────────
def fmt(n: float, sign=False) -> str:
    """원화 표시 (억/만 단위 자동 변환)"""
    abs_n = abs(n)
    sg    = ('+' if n >= 0 else '-') if sign else ('' if n >= 0 else '-')
    if abs_n >= 1_000_000_000:
        b = abs_n / 100_000_000
        return f"{sg}{b:,.1f}억원"
    elif abs_n >= 100_000_000:
        b = int(abs_n // 100_000_000)
        m = int((abs_n % 100_000_000) // 10_000)
        return f"{sg}{b}억{f' {m:,}만' if m else ''}원"
    elif abs_n >= 10_000:
        return f"{sg}{int(abs_n // 10_000):,}만원"
    else:
        return f"{sg}{int(abs_n):,}원"


# ─── 메인 앱 ──────────────────────────────────────────────────────────────────
# ── 결산 히스토리 저장 폴더 ──────────────────────────────────────────────────
FIN_HISTORY_DIR = Path(__file__).parent / "fin_history"
FIN_HISTORY_DIR.mkdir(exist_ok=True)

def main():
    # ── 비밀번호 인증 ─────────────────────────────────────────────────────────
    _PWD = st.secrets.get("app_password", "nudique0421!!")
    if "authenticated" not in st.session_state:
        st.session_state["authenticated"] = False
    if not st.session_state["authenticated"]:
        st.markdown("""
<style>
[data-testid="stAppViewContainer"] {
    background: linear-gradient(135deg, #0f2027 0%, #203a43 50%, #2c5364 100%);
}
[data-testid="stHeader"] { background: transparent; }
.login-wrap {
    max-width: 420px;
    margin: 80px auto 0;
    background: rgba(255,255,255,0.06);
    border: 1px solid rgba(255,255,255,0.12);
    border-radius: 20px;
    padding: 52px 44px 44px;
    backdrop-filter: blur(16px);
    text-align: center;
    box-shadow: 0 24px 64px rgba(0,0,0,0.4);
}
.login-logo {
    font-size: 2.6rem;
    margin-bottom: 6px;
}
.login-title {
    color: #ffffff;
    font-size: 1.45rem;
    font-weight: 800;
    letter-spacing: -0.5px;
    margin-bottom: 4px;
}
.login-sub {
    color: rgba(255,255,255,0.45);
    font-size: 0.82rem;
    margin-bottom: 36px;
    letter-spacing: 0.3px;
}
.login-divider {
    height: 1px;
    background: rgba(255,255,255,0.1);
    margin: 0 0 28px;
}
</style>
<div class="login-wrap">
  <div class="login-logo">📊</div>
  <div class="login-title">(주)비엔코 경영 Report</div>
  <div class="login-sub">Business Performance Dashboard</div>
  <div class="login-divider"></div>
</div>
""", unsafe_allow_html=True)

        _l, _c, _r = st.columns([1, 2, 1])
        with _c:
            st.markdown("<div style='height:16px'></div>", unsafe_allow_html=True)
            _pw = st.text_input("pw", type="password",
                                placeholder="🔑  비밀번호 입력",
                                label_visibility="collapsed")
            st.markdown("<div style='height:8px'></div>", unsafe_allow_html=True)
            if st.button("로그인", use_container_width=True, type="primary"):
                if _pw == _PWD:
                    st.session_state["authenticated"] = True
                    st.rerun()
                else:
                    st.error("비밀번호가 올바르지 않습니다.")
            st.markdown(
                "<p style='color:rgba(255,255,255,0.25);font-size:0.75rem;"
                "text-align:center;margin-top:20px;'>© (주)비엔코  |  Internal Use Only</p>",
                unsafe_allow_html=True)
        st.stop()
    # ── 인증 완료 ─────────────────────────────────────────────────────────────

    today      = date.today()
    total_days = calendar.monthrange(today.year, today.month)[1]
    month_label = f"{today.year}년 {today.month}월"

    # ── 데이터 로드 ──
    with st.spinner("📡 구글시트 데이터 연결 중..."):
        # 현재 월 탭 자동 탐색 — 다양한 탭명 형식 대응
        # 예: "2026.06 매출", "26.06 매출", "6월", "2026-06" 등
        _month_keywords = [
            f"{today.year}.{today.month:02d}",   # 2026.06 ← 현재 시트 형식
            f"{str(today.year)[2:]}.{today.month:02d}",  # 26.06
            f"{today.year}-{today.month:02d}",    # 2026-06
            f"{today.month}월",                   # 6월
            f"{str(today.year)[2:]}년{today.month}월",  # 26년6월
        ]
        daily_raw, daily_err = None, None
        for _kw in _month_keywords:
            daily_raw, daily_err = load_sheet_data(SHEET_IDS['daily_sales'], tab_keyword=_kw)
            if daily_raw is not None:
                break
        if daily_raw is None:
            daily_raw, daily_err = load_sheet_data(SHEET_IDS['daily_sales'])
        pl_raw,    pl_err    = load_sheet_data(SHEET_IDS['historical_pl'], by_gid=916327002)

    sales  = parse_daily_sales(daily_raw)
    detail = parse_sales_detail(daily_raw)
    avgs   = calculate_averages(pl_raw)

    elapsed_days = sales['days_with_data']
    actual_rev   = sales['total']
    daily_avg    = actual_rev / elapsed_days if elapsed_days > 0 else 0

    # ── 결산 히스토리 로드 & 직전 3개월 평균 자동 계산 ──────────────────────────
    _fin_hist   = load_fin_history()
    _fin_months = sorted(_fin_hist.keys())
    _fin_ver    = tuple(_fin_months)

    def _fgh(m, k): return _fin_hist[m]["data"].get(k, 0.0)
    def _favg_rate(key):
        rates = [_fgh(m, key) / _fgh(m, "매출액") * 100
                 for m in _fin_months[-3:] if _fgh(m, "매출액") > 0]
        return round(sum(rates) / len(rates), 1) if rates else None
    def _favg_fix(key):
        vals = [_fgh(m, key) for m in _fin_months[-3:] if _fgh(m, key) > 0]
        return float(round(sum(vals) / len(vals))) if vals else None
    def _favg_salary():
        vals = [_fgh(m,"직원급여") + _fgh(m,"상여금") + _fgh(m,"퇴직급여")
                for m in _fin_months[-3:]]
        vals = [v for v in vals if v > 0]
        return float(round(sum(vals) / len(vals))) if vals else None

    def _d(fin_val, sheet_val):
        return fin_val if fin_val is not None else float(sheet_val)

    _PARAM_DEFAULTS = {
        'p_target':  500_000_000.0,
        'p_slider':  500_000_000,
        'p_cogs':    _d(_favg_rate("매출원가"),   33.0),
        'p_adv':     _d(_favg_rate("광고선전비"), 20.0),
        'p_comm':    _d(_favg_rate("판매수수료"), 10.0),
        'p_dlv':     _d(_favg_rate("운반비"),     avgs['delivery_rate']),
        'p_fee':     _d(_favg_rate("지급수수료"), avgs['fee_rate']),
        'p_exp':     _d(_favg_rate("수출제비용"), avgs['export_rate']),
        'p_salary':  _d(_favg_salary(),           avgs['fixed']['인건비']),
        'p_welfare': _d(_favg_fix("복리후생비"),  avgs['fixed']['복리후생비']),
        'p_travel':  _d(_favg_fix("여비교통비"),  avgs['fixed']['여비교통비']),
        'p_entmt':   _d(_favg_fix("접대비"),      avgs['fixed']['접대비']),
        'p_rent':    _d(_favg_fix("지급임차료"),  avgs['fixed']['지급임차료']),
        'p_ins':     _d(_favg_fix("보험료"),      avgs['fixed']['보험료']),
        'p_tel':     _d(_favg_fix("통신비"),      avgs['fixed']['통신비']),
        'p_dep':     _d(_favg_fix("감가상각비"),  avgs['fixed']['감가상각비']),
        'p_svc':     _d(_favg_fix("용역수수료"),  avgs['fixed']['용역수수료']),
    }

    # fin_history 버전 변경 시 → 위젯 생성 전에 파라미터 자동 갱신
    _needs_update = (st.session_state.get('_fin_hist_ver') != _fin_ver)
    for _k, _v in _PARAM_DEFAULTS.items():
        if _k not in st.session_state or (_needs_update and _k not in ('p_target', 'p_slider')):
            st.session_state[_k] = _v
    if _needs_update:
        st.session_state['_fin_hist_ver'] = _fin_ver

    def _reset_to_avg():
        """모든 파라미터를 기본값으로 초기화"""
        for _k, _v in _PARAM_DEFAULTS.items():
            st.session_state[_k] = _v

    def _on_target_input():
        """사이드바 입력 → 슬라이더 동기화"""
        v = int(st.session_state.get('p_target', 500_000_000))
        st.session_state['p_slider'] = max(100_000_000, min(v, 1_000_000_000))

    def _on_slider_change():
        """슬라이더 → 사이드바 입력 동기화"""
        st.session_state['p_target'] = float(st.session_state.get('p_slider', 500_000_000))

    # ════════════════════════════════════════════════════════════════
    #  사이드바 — 판관비 파라미터
    # ════════════════════════════════════════════════════════════════
    with st.sidebar:
        st.markdown("# 📊 비엔코 시뮬레이터")
        st.caption(f"파라미터 기준: {avgs['source']}")
        st.markdown('<div class="section-divider"></div>', unsafe_allow_html=True)

        # ── 버튼 (위에 배치) ──
        col_r, col_d = st.columns(2)
        with col_r:
            if st.button("🔄 매출 새로고침", use_container_width=True,
                         help="구글시트에서 최신 매출 데이터를 다시 불러옵니다"):
                st.cache_data.clear()
                st.rerun()
        with col_d:
            if st.button("↺ 3개월 평균 초기화", use_container_width=True,
                         help="아래 모든 값을 직전 3개월 평균으로 되돌립니다"):
                _reset_to_avg()
                st.rerun()

        st.markdown('<div class="section-divider"></div>', unsafe_allow_html=True)

        # ── 월 목표 매출 ──
        st.markdown("### 🎯 월 매출 목표")
        monthly_target = int(st.number_input(
            "월 목표 매출 (원)", min_value=0, step=10_000_000,
            key='p_target', on_change=_on_target_input,
            help="슬라이더와 자동 동기화됩니다"
        ))

        st.markdown('<div class="section-divider"></div>', unsafe_allow_html=True)

        # ── 변동비 ──
        st.markdown("### 📈 변동비 (매출액 기준 %)")

        # 매출원가
        cogs_avg = avgs['cogs_rate']
        cogs_rate = st.number_input(
            "매출원가율 (%)", min_value=0.0, max_value=100.0,
            step=0.5, format="%.1f", key='p_cogs',
        )

        # 광고선전비
        adv_avg = avgs['adv_rate']
        adv_rate = st.number_input(
            "광고선전비율 (%)", min_value=0.0, max_value=100.0,
            step=0.5, format="%.1f", key='p_adv',
        )

        # 판매수수료
        comm_avg = avgs['commission_rate']
        comm_rate = st.number_input(
            "판매수수료율 (%)", min_value=0.0, max_value=100.0,
            step=0.5, format="%.1f", key='p_comm',
        )

        # 기타 변동비
        with st.expander("▸ 기타 변동비 상세 (운반비·수수료·수출)"):
            delivery_rate = st.number_input(
                "운반비율 (%)", min_value=0.0, max_value=50.0,
                step=0.5, format="%.1f", key='p_dlv',
            )
            fee_rate = st.number_input(
                "지급수수료율 (%)", min_value=0.0, max_value=30.0,
                step=0.1, format="%.1f", key='p_fee',
            )
            export_rate = st.number_input(
                "수출제비용율 (%)", min_value=0.0, max_value=20.0,
                step=0.1, format="%.1f", key='p_exp',
            )

        total_var_rate = cogs_rate + adv_rate + comm_rate + delivery_rate + fee_rate + export_rate
        gross_rate = 100.0 - cogs_rate
        net_var = adv_rate + comm_rate + delivery_rate + fee_rate + export_rate
        color_var = "normal" if net_var < gross_rate else "inverse"
        st.info(f"총 변동비율: **{total_var_rate:.1f}%** (원가 포함)  \n"
                f"매출총이익률: {gross_rate:.1f}%  /  변동판관비: {net_var:.1f}%")

        st.markdown('<div class="section-divider"></div>', unsafe_allow_html=True)

        # ── 고정비 ──
        st.markdown("### 🏢 고정비 (월 기준, 원)")
        fa = avgs['fixed']

        fixed_salary  = int(st.number_input(
            f"인건비 (급여+퇴직)  평균 {fa['인건비']:,.0f}",
            min_value=0, step=100_000, key='p_salary',
        ))
        fixed_welfare = int(st.number_input(
            f"복리후생비  평균 {fa['복리후생비']:,.0f}",
            min_value=0, step=100_000, key='p_welfare',
        ))
        fixed_travel  = int(st.number_input(
            f"여비교통비  평균 {fa['여비교통비']:,.0f}",
            min_value=0, step=50_000, key='p_travel',
        ))
        fixed_entmt   = int(st.number_input(
            f"접대비  평균 {fa['접대비']:,.0f}",
            min_value=0, step=50_000, key='p_entmt',
        ))
        fixed_rent    = int(st.number_input(
            f"지급임차료  평균 {fa['지급임차료']:,.0f}",
            min_value=0, step=10_000, key='p_rent',
        ))
        fixed_ins     = int(st.number_input(
            f"보험료  평균 {fa['보험료']:,.0f}",
            min_value=0, step=10_000, key='p_ins',
        ))
        fixed_tel     = int(st.number_input(
            f"통신비  평균 {fa['통신비']:,.0f}",
            min_value=0, step=10_000, key='p_tel',
        ))
        fixed_dep     = int(st.number_input(
            f"감가상각비  평균 {fa['감가상각비']:,.0f}",
            min_value=0, step=10_000, key='p_dep',
        ))
        fixed_svc     = int(st.number_input(
            f"용역수수료  평균 {fa['용역수수료']:,.0f}",
            min_value=0, step=10_000, key='p_svc',
        ))

        fixed_total = (fixed_salary + fixed_welfare + fixed_travel + fixed_entmt +
                       fixed_rent + fixed_ins + fixed_tel + fixed_dep + fixed_svc)
        fixed_avg_total = sum(fa.values())
        fixed_diff = fixed_total - fixed_avg_total

        st.success(f"월 고정비 합계: **{fmt(fixed_total)}**")
        if abs(fixed_diff) > 10_000:
            diff_sign = "▲" if fixed_diff > 0 else "▼"
            st.caption(f"평균 대비 {diff_sign} {fmt(abs(fixed_diff))}")

        st.markdown('<div class="section-divider"></div>', unsafe_allow_html=True)

        # ── 연결 상태 ──
        gc_status = "🟢 연결됨" if get_gspread_client() else "🔴 미연결 (데모 모드)"
        st.caption(f"Google Sheets: {gc_status}")
        if _GC_ERROR:
            st.error(f"연결 오류: {_GC_ERROR}")
        if daily_err:
            st.warning(f"매출 시트 오류: {daily_err[:60]}")
        if pl_err:
            st.warning(f"손익 시트 오류: {pl_err[:60]}")

    # ════════════════════════════════════════════════════════════════
    #  메인 영역
    # ════════════════════════════════════════════════════════════════

    # 타이틀
    c1, c2 = st.columns([4, 1])
    with c1:
        st.title(f"📊 (주)비엔코 영업이익 시뮬레이터")
    with c2:
        st.markdown(f"<br><small style='color:#6c757d;'>🕐 {datetime.now().strftime('%Y-%m-%d %H:%M')} 기준</small>",
                    unsafe_allow_html=True)
    st.caption(f"📡 매출 데이터: {sales['source']}  |  경과 **{elapsed_days}일** / {total_days}일")

    st.markdown('<div class="section-divider"></div>', unsafe_allow_html=True)

    # ── 공헌이익률 / BEP 사전 계산 ──
    var_rate_total    = (cogs_rate + adv_rate + comm_rate + delivery_rate + fee_rate + export_rate) / 100
    contribution_rate = max(1 - var_rate_total, 0.001)
    bep_rev           = fixed_total / contribution_rate

    # ── 슬라이더 (KPI보다 먼저 실행해야 projected_rev 사용 가능) ──
    st.subheader(f"🎯 {month_label} 예상 매출 설정")
    col_sl, col_info = st.columns([3, 1])
    with col_sl:
        projected_rev = st.slider(
            "월 목표",
            min_value=100_000_000,
            max_value=1_000_000_000,
            step=10_000_000,
            format="%d",
            key='p_slider',
            on_change=_on_slider_change,
            help="슬라이더를 조정하면 왼쪽 목표값과 자동 동기화됩니다",
        )
        st.caption(f"▸ 설정값: **{projected_rev:,.0f}원**")

    # 목표 기준 영업이익 (슬라이더/월목표 기준)
    proj_pl      = calc_pl(projected_rev, cogs_rate, adv_rate, comm_rate,
                           delivery_rate, fee_rate, export_rate, fixed_total)
    proj_profit  = proj_pl['op_profit']
    proj_p_color = "#198754" if proj_profit >= 0 else "#dc3545"
    proj_p_sign  = "+" if proj_profit >= 0 else ""

    # 누계 실적 기준 영업이익 (actual_rev 기준) ← KPI 큰 박스용
    actual_pl      = calc_pl(actual_rev, cogs_rate, adv_rate, comm_rate,
                             delivery_rate, fee_rate, export_rate, fixed_total)
    actual_profit  = actual_pl['op_profit']
    actual_p_color = "#198754" if actual_profit >= 0 else "#dc3545"
    actual_p_sign  = "+" if actual_profit >= 0 else ""

    with col_info:
        st.markdown(
            f"<div style='font-size:0.82rem;font-weight:700;color:#495057;"
            f"margin-bottom:6px;'>📋 실적 및 목표</div>",
            unsafe_allow_html=True
        )
        st.markdown(f"""
| 항목 | 금액 |
|------|------|
| 실적 누계 | {actual_rev:,.0f}원 |
| 월 목표 | {projected_rev:,.0f}원 |
| 목표 영업이익 | <span style='color:{proj_p_color};font-weight:700'>{proj_p_sign}{proj_profit:,.0f}원</span> |
""", unsafe_allow_html=True)

    pct_done = actual_rev / projected_rev * 100 if projected_rev > 0 else 0
    st.progress(min(pct_done / 100, 1.0),
                text=f"달성률 {pct_done:.1f}% — {actual_rev:,.0f}원 / {projected_rev:,.0f}원")

    st.markdown('<div class="section-divider"></div>', unsafe_allow_html=True)

    # ── KPI 카드 (슬라이더 이후에 배치 → projected_rev 사용 가능) ──
    remaining_days = total_days - elapsed_days
    remaining_rev  = max(projected_rev - actual_rev, 0)

    k1, k2, k3, k4 = st.columns([1, 1, 1, 2])
    with k1:
        st.metric("📈 이번달 누계 매출", f"{actual_rev:,.0f}원", f"{elapsed_days}일 확정")
    with k2:
        st.metric("📅 일평균 매출",
                  f"{daily_avg:,.0f}원" if elapsed_days > 0 else "데이터 없음")
    with k3:
        st.metric("⏳ 잔여 예상",
                  f"{remaining_rev:,.0f}원",
                  f"잔여 {remaining_days}일")
    with k4:
        # 누계 실적 기준 영업이익 — 크게 표시
        st.markdown(
            f"<div style='background:#f8f9fa;border-radius:10px;padding:14px 20px;"
            f"border-left:5px solid {actual_p_color};'>"
            f"<div style='font-size:0.8rem;color:#6c757d;font-weight:600;'>📊 누계 실적 기준 영업이익"
            f"<span style='font-size:0.72rem;margin-left:6px;'>({actual_rev:,.0f}원 기준)</span></div>"
            f"<div style='font-size:2rem;font-weight:800;color:{actual_p_color};margin-top:4px;'>"
            f"{actual_p_sign}{actual_profit:,.0f}원</div>"
            f"<div style='font-size:0.8rem;color:#6c757d;'>영업이익률 {actual_p_sign}{actual_pl['op_rate']:.1f}%</div>"
            f"</div>",
            unsafe_allow_html=True
        )

    st.markdown('<div class="section-divider"></div>', unsafe_allow_html=True)

    # ── 시나리오 계산 ──
    sc_revs = {
        '보수적':  projected_rev * 0.70,
        '기준':    projected_rev * 1.00,
        '낙관적':  projected_rev * 1.30,
    }
    scenarios = {
        name: calc_pl(rev, cogs_rate, adv_rate, comm_rate,
                      delivery_rate, fee_rate, export_rate, fixed_total)
        for name, rev in sc_revs.items()
    }
    sc_colors  = {'보수적': '#ffc107', '기준': '#0d6efd', '낙관적': '#198754'}
    sc_descs   = {'보수적': '현재 추세 ×0.7', '기준': '현재 추세 유지', '낙관적': '현재 추세 ×1.3'}

    # ── 시나리오 카드 ──
    st.subheader("📋 시나리오별 영업이익 예측")
    sc_cols = st.columns(3)
    for col, (name, r) in zip(sc_cols, scenarios.items()):
        profit = r['op_profit']
        p_color = '#198754' if profit >= 0 else '#dc3545'
        p_sign  = '+' if profit >= 0 else ''
        featured = 'featured' if name == '기준' else ''
        col.markdown(f"""
<div class="scenario-box {featured}">
  <div style="color:{sc_colors[name]};font-weight:700;font-size:0.85rem;">
    {'🟡' if name=='보수적' else '🔵' if name=='기준' else '🟢'} {name} 시나리오
  </div>
  <div style="color:#6c757d;font-size:0.78rem;margin-bottom:6px;">{sc_descs[name]}</div>
  <div style="font-size:1.0rem;font-weight:600;">매출 {fmt(r['revenue'])}</div>
  <div style="font-size:1.35rem;font-weight:700;color:{p_color};">{p_sign}{fmt(profit)}</div>
  <div style="font-size:0.78rem;color:#6c757d;">영업이익률 {p_sign}{r['op_rate']:.1f}%</div>
  <hr style="margin:8px 0; border-color:#dee2e6;">
  <div style="font-size:0.75rem;color:#6c757d;line-height:1.7;">
    매출원가 {fmt(r['cogs'])} ({cogs_rate:.1f}%)<br>
    광고비 {fmt(r['adv'])} ({adv_rate:.1f}%)<br>
    판매수수료 {fmt(r['comm'])} ({comm_rate:.1f}%)<br>
    고정비 {fmt(r['fixed_sga'])}
  </div>
</div>
""", unsafe_allow_html=True)

    st.markdown('<div class="section-divider"></div>', unsafe_allow_html=True)

    # ── 탭: 차트 ──
    tab1, tab2, tab3, tab4, tab5, tab6 = st.tabs(["📊 일별 매출", "📉 손익 구조", "📈 연도별 비교", "📂 손익계산서 분석", "🔎 역대 분석", "🔧 Google Sheets 설정"])

    # ── Tab 1: 일별 매출 ──
    with tab1:
        daily = sales['daily']
        if daily:
            days_axis  = [f"{today.month}/{i+1}" for i in range(len(daily))]
            cumulative = [sum(daily[:i+1]) for i in range(len(daily))]

            df_chart = pd.DataFrame({
                '날짜':   days_axis,
                '일별매출': daily,
                '누적매출': cumulative,
            })

            # 만원 단위로 변환해서 표시
            df_chart['일별(만원)'] = df_chart['일별매출'] / 10_000
            df_chart['누적(만원)'] = df_chart['누적매출'] / 10_000
            df_chart['레이블'] = df_chart['일별(만원)'].apply(lambda x: f"{x:,.0f}만")

            base_c = alt.Chart(df_chart).encode(
                x=alt.X('날짜:N', sort=None, title='날짜')
            )
            bars = base_c.mark_bar(color='#4e79a7', opacity=0.85).encode(
                y=alt.Y('일별(만원):Q', title='일별 매출 (만원)', axis=alt.Axis(format=',.0f')),
                tooltip=[alt.Tooltip('날짜:N'), alt.Tooltip('일별(만원):Q', title='일별매출(만원)', format=',.0f')]
            )
            bar_labels = base_c.mark_text(
                align='center', baseline='bottom', dy=-3, fontSize=10, color='#333'
            ).encode(
                y=alt.Y('일별(만원):Q'),
                text=alt.Text('레이블:N')
            )
            daily_avg_val = (actual_rev / elapsed_days / 10_000) if elapsed_days else 0
            daily_avg_line = alt.Chart(
                pd.DataFrame({'y': [daily_avg_val]})
            ).mark_rule(color='gray', strokeDash=[4, 4], opacity=0.6).encode(y='y:Q')

            line = base_c.mark_line(color='#e15759', strokeWidth=2.5, point=True).encode(
                y=alt.Y('누적(만원):Q', title='누적 매출 (만원)', axis=alt.Axis(format=',.0f')),
                tooltip=[alt.Tooltip('날짜:N'), alt.Tooltip('누적(만원):Q', title='누적(만원)', format=',.0f')]
            )
            line_labels = base_c.mark_text(
                align='left', dx=5, dy=-8, fontSize=9, color='#e15759'
            ).encode(
                y=alt.Y('누적(만원):Q'),
                text=alt.Text('누적(만원):Q', format=',.0f')
            )

            chart1 = alt.layer(bars, bar_labels, daily_avg_line).properties(
                title=f"{month_label} 일별 매출 현황 (실적 {elapsed_days}일)", height=380
            )
            chart2 = alt.layer(line, line_labels).properties(title='누적 매출 추이 (만원)', height=220)

            st.altair_chart(chart1, use_container_width=True)
            st.altair_chart(chart2, use_container_width=True)

            # 일별 테이블
            with st.expander("▸ 일별 상세 데이터"):
                df_daily = pd.DataFrame({
                    '날짜':    days_axis,
                    '매출액':  [f"{v:,.0f}" for v in daily],
                    '누적':    [f"{v:,.0f}" for v in cumulative],
                    '비중 (%)': [f"{v/actual_rev*100:.1f}%" if actual_rev > 0 else "0%" for v in daily],
                })
                st.dataframe(df_daily, use_container_width=True, hide_index=True)

            # ── 매출 상세 분석 ──────────────────────────────────────────────
            if detail:
                bc_total = detail['b2c_total'] + detail['b2b_total']

                st.markdown("---")
                st.subheader("🔍 매출 상세 분석")

                # ─────────────────────────────────────────────────────────────
                # 전체 채널 개요: B2C vs B2B 도넛
                # ─────────────────────────────────────────────────────────────
                ov_a, ov_b = st.columns([1, 2])
                with ov_a:
                    df_bc = pd.DataFrame({
                        '구분': ['B2C', 'B2B'],
                        '매출(만원)': [detail['b2c_total']/10000, detail['b2b_total']/10000],
                    })
                    donut_bc = alt.Chart(df_bc).mark_arc(innerRadius=60).encode(
                        theta=alt.Theta('매출(만원):Q'),
                        color=alt.Color('구분:N', scale=alt.Scale(
                            domain=['B2C','B2B'], range=['#4e79a7','#f28e2b'])),
                        tooltip=['구분:N', alt.Tooltip('매출(만원):Q', format=',.0f')]
                    ).properties(title='B2C vs B2B 비중', height=240)
                    st.altair_chart(donut_bc, use_container_width=True)
                with ov_b:
                    st.markdown("<br>", unsafe_allow_html=True)
                    col_bc1, col_bc2 = st.columns(2)
                    with col_bc1:
                        st.metric("🛒 B2C 누계",
                                  f"{detail['b2c_total']:,.0f}원",
                                  f"{detail['b2c_total']/bc_total*100:.1f}%" if bc_total else "")
                    with col_bc2:
                        st.metric("🏢 B2B 누계",
                                  f"{detail['b2b_total']:,.0f}원",
                                  f"{detail['b2b_total']/bc_total*100:.1f}%" if bc_total else "")

                # ═════════════════════════════════════════════════════════════
                #  B2C 영역
                # ═════════════════════════════════════════════════════════════
                st.markdown("""
<div style='background:#e8f4fd;border-left:5px solid #4e79a7;
            padding:10px 16px;border-radius:6px;margin:16px 0 10px 0;'>
  <span style='font-size:1.1rem;font-weight:700;color:#1a5276;'>🛒 B2C 영역</span>
  &nbsp;<span style='font-size:0.85rem;color:#2e86c1;'>브랜드 · 채널 · 국가별 상세</span>
</div>""", unsafe_allow_html=True)

                # ① 브랜드별
                ca, cb = st.columns(2)
                with ca:
                    df_br = pd.DataFrame([
                        {'브랜드': k, '매출(만원)': v/10000}
                        for k, v in detail['brands'].items() if v > 0
                    ])
                    if not df_br.empty:
                        bar_br = alt.Chart(df_br).mark_bar().encode(
                            x=alt.X('매출(만원):Q', title='매출 (만원)', axis=alt.Axis(format=',.0f')),
                            y=alt.Y('브랜드:N', sort='-x'),
                            color=alt.Color('브랜드:N', scale=alt.Scale(
                                range=['#59a14f','#edc948','#b07aa1'])),
                            tooltip=['브랜드:N', alt.Tooltip('매출(만원):Q', format=',.0f')]
                        ).properties(title='브랜드별 B2C 매출', height=220)
                        txt_br = bar_br.mark_text(align='left', dx=4, fontSize=10).encode(
                            text=alt.Text('매출(만원):Q', format=',.0f')
                        )
                        st.altair_chart(bar_br + txt_br, use_container_width=True)

                # ② 국가별 B2C
                with cb:
                    df_cn = pd.DataFrame([
                        {'국가': k, '매출(만원)': v/10000}
                        for k, v in detail['countries'].items() if v > 0
                    ])
                    if not df_cn.empty:
                        bar_cn = alt.Chart(df_cn).mark_bar().encode(
                            x=alt.X('매출(만원):Q', title='매출 (만원)',
                                    axis=alt.Axis(format=',.0f')),
                            y=alt.Y('국가:N', sort='-x'),
                            color=alt.Color('국가:N', scale=alt.Scale(
                                range=['#4e79a7','#f28e2b','#e15759','#76b7b2'])),
                            tooltip=['국가:N', alt.Tooltip('매출(만원):Q', format=',.0f')]
                        ).properties(title='국가별 B2C 매출', height=220)
                        txt_cn = bar_cn.mark_text(align='left', dx=4, fontSize=10).encode(
                            text=alt.Text('매출(만원):Q', format=',.0f')
                        )
                        st.altair_chart(bar_cn + txt_cn, use_container_width=True)

                # ③ 채널별 (브랜드 구분 스택)
                rows_ch = []
                for brand, chs in detail['channels_by_brand'].items():
                    for ch, v in chs.items():
                        if v > 0:
                            rows_ch.append({'채널': ch, '브랜드': brand, '매출(만원)': v/10000})
                if rows_ch:
                    df_ch = pd.DataFrame(rows_ch)
                    bar_ch = alt.Chart(df_ch).mark_bar().encode(
                        x=alt.X('sum(매출(만원)):Q', title='매출 (만원)',
                                axis=alt.Axis(format=',.0f')),
                        y=alt.Y('채널:N', sort='-x'),
                        color=alt.Color('브랜드:N', scale=alt.Scale(
                            range=['#59a14f','#edc948','#b07aa1'])),
                        tooltip=['채널:N', '브랜드:N',
                                 alt.Tooltip('매출(만원):Q', format=',.0f')]
                    ).properties(title='채널별 B2C 매출 (브랜드 구분)', height=240)
                    st.altair_chart(bar_ch, use_container_width=True)

                # ④ B2C 종합 테이블
                with st.expander("▸ B2C 채널·브랜드별 종합 표"):
                    rows_tbl = []
                    for brand, chs in detail['channels_by_brand'].items():
                        for ch, v in chs.items():
                            if v > 0:
                                b2c_sum = detail['brands'].get(brand, 1) or 1
                                rows_tbl.append({
                                    '브랜드': brand, '채널': ch,
                                    '매출액': f"{v:,.0f}원",
                                    '브랜드내비중': f"{v/b2c_sum*100:.1f}%",
                                    '전체비중': f"{v/actual_rev*100:.1f}%" if actual_rev else '-',
                                })
                        rows_tbl.append({
                            '브랜드': f'▶ {brand} 소계', '채널': '',
                            '매출액': f"{detail['brands'].get(brand,0):,.0f}원",
                            '브랜드내비중': '100%',
                            '전체비중': f"{detail['brands'].get(brand,0)/actual_rev*100:.1f}%" if actual_rev else '-',
                        })
                    rows_tbl.append({
                        '브랜드': '■ B2C 합계', '채널': '',
                        '매출액': f"{detail['b2c_total']:,.0f}원",
                        '브랜드내비중': '-',
                        '전체비중': f"{detail['b2c_total']/actual_rev*100:.1f}%" if actual_rev else '-',
                    })
                    st.dataframe(pd.DataFrame(rows_tbl), use_container_width=True, hide_index=True)

                # ═════════════════════════════════════════════════════════════
                #  B2B 영역
                # ═════════════════════════════════════════════════════════════
                st.markdown("""
<div style='background:#fef9ec;border-left:5px solid #f28e2b;
            padding:10px 16px;border-radius:6px;margin:20px 0 10px 0;'>
  <span style='font-size:1.1rem;font-weight:700;color:#784212;'>🏢 B2B 영역</span>
  &nbsp;<span style='font-size:0.85rem;color:#b7770d;'>국가 · 브랜드 · 벤더별 상세</span>
</div>""", unsafe_allow_html=True)

                # B2B 국가별 / 브랜드별 / 벤더 상세
                b2b_details   = detail.get('b2b_details', [])
                b2b_countries = detail.get('b2b_countries', {})
                b2b_brands    = detail.get('b2b_brands', {})

                # 국가 색상 팔레트 (다국가 대응)
                COUNTRY_COLORS = [
                    '#f28e2b','#4e79a7','#e15759','#76b7b2',
                    '#59a14f','#edc948','#b07aa1','#ff9da7',
                    '#9c755f','#bab0ac','#d62728','#aec7e8',
                ]

                if b2b_details:
                    be_a, be_b = st.columns(2)

                    # ① 국가별 B2B 가로 바 차트
                    with be_a:
                        df_b2b_cn = pd.DataFrame(
                            sorted(
                                [{'국가': k, '매출(만원)': v/10000}
                                 for k, v in b2b_countries.items() if v > 0],
                                key=lambda x: -x['매출(만원)']
                            )
                        )
                        if not df_b2b_cn.empty:
                            n_cn = len(df_b2b_cn)
                            bar_b2b_cn = alt.Chart(df_b2b_cn).mark_bar().encode(
                                x=alt.X('매출(만원):Q', title='매출 (만원)',
                                        axis=alt.Axis(format=',.0f')),
                                y=alt.Y('국가:N', sort='-x', title='국가'),
                                color=alt.Color('국가:N',
                                    scale=alt.Scale(range=COUNTRY_COLORS[:max(n_cn, 4)]),
                                    legend=None),
                                tooltip=['국가:N',
                                         alt.Tooltip('매출(만원):Q', format=',.0f', title='매출(만원)')]
                            ).properties(
                                title='국가별 B2B 매출',
                                height=max(200, n_cn * 36)
                            )
                            txt_b2b_cn = bar_b2b_cn.mark_text(
                                align='left', dx=4, fontSize=10
                            ).encode(text=alt.Text('매출(만원):Q', format=',.0f'))
                            st.altair_chart(bar_b2b_cn + txt_b2b_cn, use_container_width=True)

                    # ② 브랜드별 B2B 가로 바 차트
                    with be_b:
                        df_b2b_br = pd.DataFrame([
                            {'브랜드': k, '매출(만원)': v/10000}
                            for k, v in b2b_brands.items() if v > 0
                        ])
                        if not df_b2b_br.empty:
                            bar_b2b_br = alt.Chart(df_b2b_br).mark_bar().encode(
                                x=alt.X('매출(만원):Q', title='매출 (만원)',
                                        axis=alt.Axis(format=',.0f')),
                                y=alt.Y('브랜드:N', sort='-x'),
                                color=alt.Color('브랜드:N', scale=alt.Scale(
                                    range=['#59a14f','#edc948','#b07aa1'])),
                                tooltip=['브랜드:N',
                                         alt.Tooltip('매출(만원):Q', format=',.0f')]
                            ).properties(title='브랜드별 B2B 매출', height=220)
                            txt_b2b_br = bar_b2b_br.mark_text(
                                align='left', dx=4, fontSize=10
                            ).encode(text=alt.Text('매출(만원):Q', format=',.0f'))
                            st.altair_chart(bar_b2b_br + txt_b2b_br, use_container_width=True)

                    # ③ 벤더별 도넛 차트 (매출 있는 벤더만)
                    df_vendor = pd.DataFrame([
                        {'벤더': d['label'], '국가': d['country'],
                         '브랜드': d['brand'], '매출(만원)': d['value'] / 10000}
                        for d in b2b_details if d['value'] > 0
                    ])
                    if not df_vendor.empty:
                        st.markdown("**🏷️ 벤더별 B2B 매출 비중**")
                        v_a, v_b = st.columns([1, 1])
                        with v_a:
                            # 도넛 — 벤더 레이블이 길 수 있으므로 상위 8개만 명시
                            donut_vendor = alt.Chart(df_vendor).mark_arc(
                                innerRadius=55, outerRadius=110
                            ).encode(
                                theta=alt.Theta('매출(만원):Q', stack=True),
                                color=alt.Color('벤더:N',
                                    scale=alt.Scale(scheme='tableau20'),
                                    legend=alt.Legend(title='벤더', orient='right',
                                                      labelLimit=160)),
                                tooltip=['벤더:N', '국가:N', '브랜드:N',
                                         alt.Tooltip('매출(만원):Q', format=',.0f')]
                            ).properties(height=320)
                            st.altair_chart(donut_vendor, use_container_width=True)

                        with v_b:
                            # 벤더 요약 테이블
                            b2b_tot = detail['b2b_total'] or 1
                            df_vendor['비중(%)'] = (
                                df_vendor['매출(만원)'] * 10000 / b2b_tot * 100
                            ).round(1)
                            df_vendor['매출액'] = df_vendor['매출(만원)'].apply(
                                lambda x: f"{x*10000:,.0f}원"
                            )
                            st.dataframe(
                                df_vendor[['국가','브랜드','벤더','매출액','비중(%)']].sort_values(
                                    '비중(%)', ascending=False
                                ).reset_index(drop=True),
                                use_container_width=True, hide_index=True
                            )
                else:
                    # B2B 상세 행 못 찾은 경우
                    st.info(
                        f"📌 B2B 누계 매출: **{detail['b2b_total']:,.0f}원**  \n"
                        "상세 벤더·국가·브랜드 데이터를 시트에서 불러오는 중입니다."
                    )

        else:
            st.info("일별 매출 데이터를 불러오지 못했습니다. 사이드바의 연결 상태를 확인하세요.")

    # ── Tab 2: 손익 구조 ──
    with tab2:
        base = scenarios['기준']

        # 손익 구조 바 차트 (워터폴 대체)
        wf_items = [
            ('매출액',     base['revenue'],              'total'),
            ('매출원가',   -base['cogs'],                'negative'),
            ('매출총이익', base['gross'],                 'total'),
            ('광고비',     -base['adv'],                 'negative'),
            ('판매수수료', -base['comm'],                'negative'),
            ('운반비',     -base['dlv'],                 'negative'),
            ('기타변동비', -(base['fee']+base['exp']),   'negative'),
            ('고정비',     -base['fixed_sga'],           'negative'),
            ('영업이익',   base['op_profit'],
             'profit' if base['op_profit'] >= 0 else 'loss'),
        ]
        wf_df = pd.DataFrame(wf_items, columns=['항목', '금액(만원)', '유형'])
        wf_df['금액(만원)'] = wf_df['금액(만원)'] / 10_000

        color_scale = alt.Scale(
            domain=['total', 'negative', 'profit', 'loss'],
            range=['#0d6efd',  '#dc3545',  '#198754', '#c0392b']
        )
        chart_wf = alt.Chart(wf_df).mark_bar(size=40).encode(
            x=alt.X('항목:N', sort=None, title=''),
            y=alt.Y('금액(만원):Q', title='금액 (만원)', axis=alt.Axis(format=',.0f')),
            color=alt.Color('유형:N', scale=color_scale, legend=None),
            tooltip=[alt.Tooltip('항목:N'), alt.Tooltip('금액(만원):Q', format=',.0f', title='금액(만원)')]
        ).properties(
            title=f"손익 구조 — 기준 시나리오 (매출 {fmt(base['revenue'])})",
            height=400
        )
        zero_rule = alt.Chart(pd.DataFrame({'y': [0]})).mark_rule(
            color='gray', strokeDash=[4, 4], opacity=0.5
        ).encode(y='y:Q')
        st.altair_chart(chart_wf + zero_rule, use_container_width=True)

        # 비용 구성 도넛 + 요약 테이블
        col_d1, col_d2 = st.columns(2)
        with col_d1:
            pie_df = pd.DataFrame({
                '항목':  ['매출원가', '광고선전비', '판매수수료', '운반비', '기타변동비', '고정비'],
                '금액':  [base['cogs'], base['adv'], base['comm'], base['dlv'],
                          base['fee']+base['exp'], base['fixed_sga']],
            })
            pie_df['비율'] = (pie_df['금액'] / pie_df['금액'].sum() * 100).round(1)
            donut = alt.Chart(pie_df).mark_arc(innerRadius=50).encode(
                theta=alt.Theta('금액:Q'),
                color=alt.Color('항목:N', scale=alt.Scale(
                    range=['#dc3545','#ffc107','#fd7e14','#0d6efd','#6f42c1','#20c997']
                )),
                tooltip=['항목:N', alt.Tooltip('금액:Q', format=',.0f'), alt.Tooltip('비율:Q', format='.1f', title='%')]
            ).properties(title='비용 구성 (기준 시나리오)', height=300)
            st.altair_chart(donut, use_container_width=True)

        with col_d2:
            st.markdown("**시나리오별 손익 요약 (단위: 만원)**")
            summary = []
            for name, r in scenarios.items():
                profit = r['op_profit']
                summary.append({
                    '시나리오':   name,
                    '매출액':     f"{r['revenue']/10000:,.0f}",
                    '매출총이익': f"{r['gross']/10000:,.0f}",
                    '변동판관비': f"{r['var_sga']/10000:,.0f}",
                    '고정판관비': f"{r['fixed_sga']/10000:,.0f}",
                    '영업이익':   f"{'+' if profit >= 0 else ''}{profit/10000:,.0f}",
                    '이익률':     f"{'+' if r['op_rate'] >= 0 else ''}{r['op_rate']:.1f}%",
                })
            st.dataframe(pd.DataFrame(summary), use_container_width=True, hide_index=True)

    # ── Tab 3: 연도별 비교 ──
    with tab3:
        periods  = ['2024 월평균', '2025 월평균', '2026년 1월', '2026년 2월',
                    '2026년 3월', '2026년 4월', f'{month_label} (예측)']
        revenues = [235_834_533, 323_486_830, 286_493_155, 231_348_650,
                    248_143_781, 361_297_124, projected_rev]
        profits  = [-166_992_490, -100_110_601, None, None,
                    None, 29_393_517, scenarios['기준']['op_profit']]

        # 매출 바 차트
        rev_df = pd.DataFrame({'기간': periods, '매출액(만원)': [v/10000 for v in revenues]})
        chart_rev = alt.Chart(rev_df).mark_bar(color='#4e79a7', opacity=0.75).encode(
            x=alt.X('기간:N', sort=None, title=''),
            y=alt.Y('매출액(만원):Q', title='금액 (만원)', axis=alt.Axis(format=',.0f')),
            tooltip=[alt.Tooltip('기간:N'), alt.Tooltip('매출액(만원):Q', format=',.0f')]
        ).properties(title='월별 매출액 추이', height=280)

        # 영업이익 바 차트 (있는 것만)
        profit_rows = [(p, v/10000) for p, v in zip(periods, profits) if v is not None]
        profit_df = pd.DataFrame(profit_rows, columns=['기간', '영업이익(만원)'])
        profit_df['색상'] = profit_df['영업이익(만원)'].apply(lambda x: '흑자' if x >= 0 else '적자')

        chart_profit = alt.Chart(profit_df).mark_bar(opacity=0.9).encode(
            x=alt.X('기간:N', sort=None, title=''),
            y=alt.Y('영업이익(만원):Q', title='영업이익 (만원)', axis=alt.Axis(format=',.0f')),
            color=alt.Color('색상:N', scale=alt.Scale(
                domain=['흑자', '적자'], range=['#198754', '#dc3545']
            ), legend=alt.Legend(title='구분')),
            tooltip=[alt.Tooltip('기간:N'), alt.Tooltip('영업이익(만원):Q', format=',.0f')]
        )
        zero_line2 = alt.Chart(pd.DataFrame({'y': [0]})).mark_rule(
            color='gray', strokeDash=[4, 4], opacity=0.5
        ).encode(y='y:Q')
        chart_profit_final = (chart_profit + zero_line2).properties(
            title='월별 영업이익 추이', height=280
        )

        st.altair_chart(chart_rev, use_container_width=True)
        st.altair_chart(chart_profit_final, use_container_width=True)

    # ── Tab 4: 손익계산서 분석 (Ver2) ──
    with tab4:
        st.subheader("📂 손익계산서 누적 관리")
        st.caption("월별 결산 엑셀을 업로드하면 자동 저장됩니다. 여러 달을 쌓아 추이를 확인하세요.")

        # ── 파일 업로드 ──
        uploaded = st.file_uploader(
            "📎 월별 결산 엑셀 업로드 (.xlsx)",
            type=["xlsx"],
            help="손익(기간별) 시트가 포함된 엑셀 파일",
            key="fin_upload",
        )
        if uploaded:
            with st.spinner("파싱 중..."):
                fin = parse_financial_excel(uploaded)
            if fin is None or "error" in (fin or {}):
                st.error(f"파싱 실패: {(fin or {}).get('error','알 수 없는 오류')}")
            else:
                saved_months = []
                for m in fin["months"]:
                    item_vals = {k: (fin["data"].get(k) or {}).get(m, 0.0) for k in _FIN_ITEMS}
                    save_fin_month(m, fin["filename"], item_vals)
                    saved_months.append(m)
                st.success(f"✅ {len(saved_months)}개월 저장 완료: {', '.join(saved_months)}")
                st.rerun()



        # ── 히스토리 로드 ──
        history = load_fin_history()

        if not history:
            st.info("📎 아직 저장된 결산자료가 없습니다. 위에서 엑셀을 업로드하세요.")
        else:
            all_months = sorted(history.keys())

            # ── 저장 파일 목록 & 삭제 ──
            with st.expander(f"📁 저장된 결산자료 ({len(all_months)}개월)", expanded=False):
                for m in sorted(all_months, reverse=True):
                    rec = history[m]
                    c1, c2, c3 = st.columns([2, 3, 1])
                    c1.markdown(f"**{m}**")
                    c2.caption(f"{rec.get('filename','')}  ·  {rec.get('uploaded_at','')}")
                    if c3.button("🗑️", key=f"del_{m}", help=f"{m} 삭제"):
                        (FIN_HISTORY_DIR / f"{m}.json").unlink(missing_ok=True)
                        st.rerun()

            st.markdown("---")

            # ── 월별 손익 요약 테이블 ──
            def gh(m, key): return history[m]["data"].get(key, 0.0)

            tbl_rows = []
            for m in all_months:
                r = gh(m, "매출액"); c = gh(m, "매출원가")
                g = gh(m, "매출총이익"); s = gh(m, "판관비계"); e = gh(m, "영업손익")
                tbl_rows.append({
                    "월": m,
                    "매출액(만)":   f"{r/10000:,.0f}",
                    "매출원가율":   f"{c/r*100:.1f}%" if r else "-",
                    "매출총이익(만)": f"{g/10000:,.0f}",
                    "총이익률":    f"{g/r*100:.1f}%" if r else "-",
                    "판관비(만)":  f"{s/10000:,.0f}",
                    "판관비율":    f"{s/r*100:.1f}%" if r else "-",
                    "영업손익(만)": f"{e/10000:+,.0f}",
                    "영업이익률":  f"{e/r*100:.1f}%" if r else "-",
                })
            st.markdown("**📋 누적 월별 손익 요약**")
            st.dataframe(pd.DataFrame(tbl_rows), use_container_width=True, hide_index=True)

            st.markdown("---")

            # ── 차트 1: 매출·매출총이익 그룹바 ──
            rev_rows = []
            for m in all_months:
                r = gh(m, "매출액"); g = gh(m, "매출총이익"); e = gh(m, "영업손익")
                if r == 0: continue
                rev_rows += [
                    {"월": m, "항목": "매출액",     "금액(만)": r/10000},
                    {"월": m, "항목": "매출총이익", "금액(만)": g/10000},
                ]
            if rev_rows:
                bar_rev = alt.Chart(pd.DataFrame(rev_rows)).mark_bar(opacity=0.85).encode(
                    x=alt.X("월:N", sort=None, title=""),
                    y=alt.Y("금액(만):Q", title="금액 (만원)", axis=alt.Axis(format=",.0f")),
                    color=alt.Color("항목:N", scale=alt.Scale(
                        domain=["매출액","매출총이익"], range=["#4e79a7","#59a14f"])),
                    xOffset="항목:N",
                    tooltip=["월:N","항목:N", alt.Tooltip("금액(만):Q", format=",.0f")],
                ).properties(title="월별 매출액 & 매출총이익 (만원)", height=260)
                st.altair_chart(bar_rev, use_container_width=True)

            # ── 차트 2: 영업손익 바 (적자=빨강, 흑자=초록) ──
            ebit_rows = [{"월": m, "영업손익(만)": gh(m,"영업손익")/10000,
                          "구분": "흑자" if gh(m,"영업손익")>=0 else "적자"}
                         for m in all_months if gh(m,"매출액") > 0]
            if ebit_rows:
                df_eb = pd.DataFrame(ebit_rows)
                bar_eb = alt.Chart(df_eb).mark_bar().encode(
                    x=alt.X("월:N", sort=None, title=""),
                    y=alt.Y("영업손익(만):Q", title="영업손익 (만원)", axis=alt.Axis(format=",.0f")),
                    color=alt.Color("구분:N", scale=alt.Scale(
                        domain=["흑자","적자"], range=["#198754","#dc3545"])),
                    tooltip=["월:N", alt.Tooltip("영업손익(만):Q", format=",.1f")],
                )
                zero_r = alt.Chart(pd.DataFrame({"y":[0]})).mark_rule(
                    color="gray", strokeDash=[4,4]).encode(y="y:Q")
                st.altair_chart((bar_eb+zero_r).properties(
                    title="월별 영업손익 추이 (만원)", height=220), use_container_width=True)

            # ── 차트 3: 판관비 항목별 스택 바 ──
            SGA_ITEMS = [
                ("광고선전비","#e15759"),("판매수수료","#f28e2b"),("운반비","#edc948"),
                ("직원급여","#4e79a7"),("지급수수료","#76b7b2"),("수출제비용","#59a14f"),
                ("복리후생비","#b07aa1"),("여비교통비","#ff9da7"),("보험료","#9c755f"),
                ("감가상각비","#bab0ac"),("통신비","#aec7e8"),("지급임차료","#ffbb78"),
                ("용역수수료","#98df8a"),("소모품비","#d62728"),
            ]
            sga_rows = []
            for m in all_months:
                if gh(m,"매출액") == 0: continue
                for itm,_ in SGA_ITEMS:
                    v = gh(m, itm)
                    if v > 0:
                        sga_rows.append({"월": m, "항목": itm, "금액(만)": v/10000})
            if sga_rows:
                bar_sga = alt.Chart(pd.DataFrame(sga_rows)).mark_bar().encode(
                    x=alt.X("월:N", sort=None, title=""),
                    y=alt.Y("금액(만):Q", title="판관비 (만원)", axis=alt.Axis(format=",.0f")),
                    color=alt.Color("항목:N",
                        scale=alt.Scale(domain=[i for i,_ in SGA_ITEMS],
                                        range=[c for _,c in SGA_ITEMS]),
                        legend=alt.Legend(title="항목")),
                    tooltip=["월:N","항목:N", alt.Tooltip("금액(만):Q", format=",.0f")],
                ).properties(title="월별 판관비 항목별 구성 (만원)", height=300)
                st.altair_chart(bar_sga, use_container_width=True)

            st.markdown("---")

            # ── 직전 3개월 시뮬레이터 적용 ──
            last3 = all_months[-3:]
            st.markdown(f"**🔧 시뮬레이터 적용 예정 값 (직전 3개월: {', '.join(last3)} 평균)**")

            def _avg_rate_h(num_key):
                rates = []
                for m in last3:
                    r = gh(m,"매출액"); v = gh(m, num_key)
                    if r > 0: rates.append(v/r*100)
                return round(sum(rates)/len(rates), 1) if rates else 0.0

            def _avg_fix_h(key):
                vals = [gh(m,key) for m in last3 if gh(m,key) > 0]
                return round(sum(vals)/len(vals)) if vals else 0

            def _avg_salary_h():
                totals = [gh(m,"직원급여")+gh(m,"상여금")+gh(m,"퇴직급여") for m in last3]
                vals = [v for v in totals if v > 0]
                return round(sum(vals)/len(vals)) if vals else 0

            sim_p_h = {
                "p_cogs":    _avg_rate_h("매출원가"),
                "p_adv":     _avg_rate_h("광고선전비"),
                "p_comm":    _avg_rate_h("판매수수료"),
                "p_dlv":     _avg_rate_h("운반비"),
                "p_fee":     _avg_rate_h("지급수수료"),
                "p_exp":     _avg_rate_h("수출제비용"),
                "p_salary":  float(_avg_salary_h()),
                "p_welfare": float(_avg_fix_h("복리후생비")),
                "p_travel":  float(_avg_fix_h("여비교통비")),
                "p_entmt":   float(_avg_fix_h("접대비")),
                "p_rent":    float(_avg_fix_h("지급임차료")),
                "p_ins":     float(_avg_fix_h("보험료")),
                "p_tel":     float(_avg_fix_h("통신비")),
                "p_dep":     float(_avg_fix_h("감가상각비")),
                "p_svc":     float(_avg_fix_h("용역수수료")),
            }
            sim_p_h = {k: v for k, v in sim_p_h.items() if v > 0}

            st.info(f"📅 저장 기간: **{all_months[0]} ~ {all_months[-1]}**  |  "
                    f"사이드바 자동 반영 기준: 직전 3개월 ({', '.join(last3)})")

    # ══════════════════════════════════════════════════════════════════
    #  Tab 5: 역대 시뮬레이션 & 실적 비교
    # ══════════════════════════════════════════════════════════════════
    with tab5:
        st.subheader("🔎 역대 시뮬레이션 & 실적 비교")

        # ══════════════════════════════════════════════════════════════════
        #  Section 0: 월별 매출 상세분석 (구글시트 탭 선택)
        # ══════════════════════════════════════════════════════════════════
        st.markdown("### 📅 월별 매출 상세분석")

        _tabs_list = get_daily_sheet_tabs()
        if not _tabs_list:
            st.info("📡 구글시트 연결 후 월별 탭이 표시됩니다. (설정 탭에서 credentials.json 확인)")
        else:
            _tab_names = [t[0] for t in _tabs_list]
            _tab_gids  = {t[0]: t[1] for t in _tabs_list}
            _sel_tab   = st.selectbox(
                "분석할 월 선택 (구글시트 탭)",
                _tab_names, index=0, key="hist_month_sel",
                help="구글시트의 월별 탭을 선택하면 해당 월 일별 매출이 상세 분석됩니다."
            )
            if _sel_tab:
                with st.spinner(f"📡 {_sel_tab} 데이터 로딩 중..."):
                    _h_raw, _h_err = load_sheet_data(
                        SHEET_IDS['daily_sales'], by_gid=_tab_gids[_sel_tab]
                    )
                if _h_err:
                    st.error(f"로딩 실패: {_h_err}")
                else:
                    _hs   = parse_daily_sales(_h_raw)
                    _hd   = parse_sales_detail(_h_raw)
                    _htot = _hs['total']
                    _hdays = _hs['days_with_data']

                    # ── 요약 메트릭 ──────────────────────────────────────
                    _mc1, _mc2, _mc3 = st.columns(3)
                    _mc1.metric(f"📦 {_sel_tab} 총매출", f"{_htot:,.0f}원")
                    _mc2.metric("🛒 B2C", f"{_hs.get('b2c',0):,.0f}원")
                    _mc3.metric("🏢 B2B", f"{_hs.get('b2b',0):,.0f}원")

                    # ── 일별 매출 바차트 ─────────────────────────────────
                    _daily = _hs['daily']
                    if _daily:
                        _days_ax = [f"{i+1}일" for i in range(len(_daily))]
                        _cumul   = [sum(_daily[:i+1]) for i in range(len(_daily))]
                        _df_d = pd.DataFrame({
                            '날짜': _days_ax,
                            '일별(만원)': [v/10000 for v in _daily],
                            '누적(만원)': [v/10000 for v in _cumul],
                        })
                        _base = alt.Chart(_df_d).encode(x=alt.X('날짜:N', sort=None))
                        _bars = _base.mark_bar(color='#4e79a7', opacity=0.85).encode(
                            y=alt.Y('일별(만원):Q', title='일별 매출 (만원)',
                                    axis=alt.Axis(format=',.0f')),
                            tooltip=['날짜:N', alt.Tooltip('일별(만원):Q', format=',.0f')]
                        )
                        _line = _base.mark_line(color='#e15759', strokeWidth=2.5, point=True).encode(
                            y=alt.Y('누적(만원):Q', title='누적 (만원)',
                                    axis=alt.Axis(format=',.0f')),
                            tooltip=['날짜:N', alt.Tooltip('누적(만원):Q', format=',.0f')]
                        )
                        st.altair_chart(
                            alt.layer(_bars).properties(
                                title=f"{_sel_tab} 일별 매출 ({_hdays}일 실적)", height=300
                            ), use_container_width=True
                        )
                        st.altair_chart(
                            _line.properties(title='누적 매출 추이', height=180),
                            use_container_width=True
                        )

                    # ── 매출 상세 (B2C / B2B) ────────────────────────────
                    if _hd:
                        _bc_tot = _hd['b2c_total'] + _hd['b2b_total']
                        st.markdown("**📊 B2C vs B2B 비중**")
                        _doa, _dob = st.columns([1, 2])
                        with _doa:
                            _df_bc = pd.DataFrame({
                                '구분': ['B2C','B2B'],
                                '매출(만)': [_hd['b2c_total']/10000,
                                             _hd['b2b_total']/10000],
                            })
                            st.altair_chart(
                                alt.Chart(_df_bc).mark_arc(innerRadius=55).encode(
                                    theta='매출(만):Q',
                                    color=alt.Color('구분:N', scale=alt.Scale(
                                        domain=['B2C','B2B'], range=['#4e79a7','#f28e2b'])),
                                    tooltip=['구분:N', alt.Tooltip('매출(만):Q', format=',.0f')]
                                ).properties(height=220), use_container_width=True
                            )
                        with _dob:
                            # 브랜드별 B2C
                            _df_br = pd.DataFrame([
                                {'브랜드': k, '매출(만)': v/10000}
                                for k, v in _hd['brands'].items() if v > 0
                            ])
                            if not _df_br.empty:
                                st.altair_chart(
                                    alt.Chart(_df_br).mark_bar().encode(
                                        x=alt.X('매출(만):Q', title='매출 (만원)',
                                                axis=alt.Axis(format=',.0f')),
                                        y=alt.Y('브랜드:N', sort='-x'),
                                        color=alt.Color('브랜드:N', scale=alt.Scale(
                                            range=['#59a14f','#edc948','#b07aa1'])),
                                        tooltip=['브랜드:N',
                                                 alt.Tooltip('매출(만):Q', format=',.0f')]
                                    ).properties(title='브랜드별 B2C 매출', height=220),
                                    use_container_width=True
                                )

                        # B2B 상세
                        _b2b_det = _hd.get('b2b_details', [])
                        _b2b_cnt = _hd.get('b2b_countries', {})
                        _b2b_brd = _hd.get('b2b_brands', {})

                        if _b2b_det:
                            st.markdown("**🏢 B2B 상세**")
                            _ba, _bb = st.columns(2)
                            with _ba:
                                _df_cnt = pd.DataFrame([
                                    {'국가': k, '매출(만)': v/10000}
                                    for k, v in _b2b_cnt.items() if v > 0
                                ])
                                if not _df_cnt.empty:
                                    st.altair_chart(
                                        alt.Chart(_df_cnt).mark_bar().encode(
                                            x=alt.X('매출(만):Q', axis=alt.Axis(format=',.0f')),
                                            y=alt.Y('국가:N', sort='-x'),
                                            color=alt.Color('국가:N'),
                                            tooltip=['국가:N',
                                                     alt.Tooltip('매출(만):Q', format=',.0f')]
                                        ).properties(title='국가별 B2B 매출', height=250),
                                        use_container_width=True
                                    )
                            with _bb:
                                _df_vnd = pd.DataFrame([
                                    {'벤더': d['label'], '매출(만)': d['value']/10000}
                                    for d in _b2b_det if d.get('value', 0) > 0
                                ])
                                if not _df_vnd.empty:
                                    st.altair_chart(
                                        alt.Chart(_df_vnd).mark_arc(innerRadius=50).encode(
                                            theta='매출(만):Q',
                                            color=alt.Color('벤더:N'),
                                            tooltip=['벤더:N',
                                                     alt.Tooltip('매출(만):Q', format=',.0f')]
                                        ).properties(title='벤더별 B2B 비중', height=250),
                                        use_container_width=True
                                    )
                    st.markdown("---")

        st.caption("구글시트 월별 매출 × 현재 시뮬 파라미터 = 시뮬 손익 / 결산 업로드 자료 = 실적 손익")

        # ── 데이터 준비 ──────────────────────────────────────────────────
        sheet_pl  = parse_sheet_monthly_pl(pl_raw)   # 구글시트 월별 실적
        fin_hist  = load_fin_history()               # 결산 업로드 월별 실적
        ss        = st.session_state                 # 현재 시뮬 파라미터

        # 구글시트 OR fin_history 어느 쪽이든 있는 월 합집합
        all_months = sorted(set(sheet_pl.keys()) | set(fin_hist.keys()))

        if not all_months:
            st.info("📡 구글시트 연결 또는 결산 파일 업로드 후 이용 가능합니다.")
        else:
            # ── 월별 시뮬레이션 계산 ───────────────────────────────────────
            # 수입: 구글시트 있으면 우선, 없으면 fin_history 매출 사용
            sim_results = {}
            for m in all_months:
                rev = (sheet_pl.get(m, {}).get('매출액') or
                       fin_hist.get(m, {}).get('data', {}).get('매출액', 0))
                if rev > 0:
                    sim_results[m] = simulate_month_pl(rev, ss)

            # ── Section 1: 역대 시뮬레이션 ───────────────────────────────
            st.markdown("### 📈 역대 월별 시뮬레이션")

            sim_tbl = []
            for m in all_months:
                s = sim_results.get(m)
                if not s:
                    continue
                r = s['매출액']
                sim_tbl.append({
                    "월":          m,
                    "매출(만)":    f"{r/10000:,.0f}",
                    "시뮬원가율":  f"{s['매출원가']/r*100:.1f}%" if r else "-",
                    "시뮬총이익(만)": f"{s['매출총이익']/10000:,.0f}",
                    "시뮬판관비(만)": f"{s['판관비계']/10000:,.0f}",
                    "시뮬영업손익(만)": f"{s['영업손익']/10000:+,.0f}",
                    "시뮬영업이익률": f"{s['영업손익']/r*100:.1f}%" if r else "-",
                })
            if sim_tbl:
                st.dataframe(pd.DataFrame(sim_tbl), use_container_width=True, hide_index=True)

            # 차트: 매출 + 시뮬 영업손익
            chart_rows = []
            for m in all_months:
                s = sim_results.get(m)
                if not s:
                    continue
                chart_rows += [
                    {"월": m, "항목": "매출액",      "금액(만)": s['매출액']/10000},
                    {"월": m, "항목": "시뮬 영업손익","금액(만)": s['영업손익']/10000},
                ]
            if chart_rows:
                df_c = pd.DataFrame(chart_rows)
                bar_sim = alt.Chart(df_c).mark_bar(opacity=0.8).encode(
                    x=alt.X("월:N", sort=None, title=""),
                    y=alt.Y("금액(만):Q", title="금액 (만원)", axis=alt.Axis(format=",.0f")),
                    color=alt.Color("항목:N", scale=alt.Scale(
                        domain=["매출액","시뮬 영업손익"],
                        range=["#4e79a7","#f28e2b"])),
                    xOffset="항목:N",
                    tooltip=["월:N","항목:N", alt.Tooltip("금액(만):Q", format=",.1f")],
                ).properties(title="역대 월별 매출 & 시뮬 영업손익 (만원)", height=270)
                st.altair_chart(bar_sim, use_container_width=True)

            st.markdown("---")

            # ── Section 2: 시뮬 vs 실적 비교 (결산 업로드 월) ──────────────
            compare_months = [m for m in all_months
                              if m in sim_results and m in fin_hist]

            st.markdown("### 📊 시뮬레이션 vs 실적 비교")

            if not compare_months:
                st.info("📎 Tab 4 에서 결산 파일을 업로드하면 비교 분석이 활성화됩니다.")
            else:
                COMPARE_ITEMS = [
                    ("매출액",    "매출액"),
                    ("매출원가",  "매출원가"),
                    ("매출총이익","매출총이익"),
                    ("판관비계",  "판관비계"),
                    ("영업손익",  "영업손익"),
                ]

                # 종합 비교 테이블
                cmp_tbl = []
                for m in compare_months:
                    s   = sim_results[m]
                    act = fin_hist[m]["data"]
                    rev_act = act.get("매출액", 0)
                    for sim_key, act_key in COMPARE_ITEMS:
                        sv = s.get(sim_key, 0)
                        av = act.get(act_key, 0)
                        diff = av - sv
                        pct  = diff / abs(sv) * 100 if sv != 0 else 0
                        cmp_tbl.append({
                            "월": m, "항목": sim_key,
                            "시뮬(만)":  f"{sv/10000:+,.0f}",
                            "실적(만)":  f"{av/10000:+,.0f}",
                            "차이(만)":  f"{diff/10000:+,.0f}",
                            "차이율":    f"{pct:+.1f}%",
                        })

                st.markdown("**📋 항목별 시뮬 vs 실적 비교표**")
                st.dataframe(pd.DataFrame(cmp_tbl), use_container_width=True, hide_index=True)

                st.markdown("---")

                # 영업손익 시뮬 vs 실적 차트
                op_rows = []
                for m in compare_months:
                    s   = sim_results[m]
                    act = fin_hist[m]["data"]
                    op_rows += [
                        {"월": m, "구분": "시뮬",  "영업손익(만)": s["영업손익"]/10000},
                        {"월": m, "구분": "실적",  "영업손익(만)": act.get("영업손익",0)/10000},
                    ]
                if op_rows:
                    df_op = pd.DataFrame(op_rows)
                    bar_op = alt.Chart(df_op).mark_bar(opacity=0.85).encode(
                        x=alt.X("월:N", sort=None, title=""),
                        y=alt.Y("영업손익(만):Q", title="영업손익 (만원)",
                                axis=alt.Axis(format=",.0f")),
                        color=alt.Color("구분:N", scale=alt.Scale(
                            domain=["시뮬","실적"],
                            range=["#f28e2b","#4e79a7"])),
                        xOffset="구분:N",
                        tooltip=["월:N","구분:N",
                                 alt.Tooltip("영업손익(만):Q", format=",.1f")],
                    )
                    zero_l = alt.Chart(pd.DataFrame({"y":[0]})).mark_rule(
                        color="gray", strokeDash=[4,4]).encode(y="y:Q")
                    st.altair_chart(
                        (bar_op + zero_l).properties(
                            title="월별 영업손익: 시뮬 vs 실적 (만원)", height=260),
                        use_container_width=True)

                # 차이 분석: 영업손익 갭 바
                diff_rows = []
                for m in compare_months:
                    s   = sim_results[m]
                    act = fin_hist[m]["data"]
                    diff = (act.get("영업손익",0) - s["영업손익"]) / 10000
                    diff_rows.append({"월": m, "차이(만)": diff,
                                      "방향": "실적↑" if diff >= 0 else "실적↓"})
                if diff_rows:
                    df_diff = pd.DataFrame(diff_rows)
                    bar_diff = alt.Chart(df_diff).mark_bar().encode(
                        x=alt.X("월:N", sort=None, title=""),
                        y=alt.Y("차이(만):Q", title="실적 - 시뮬 (만원)",
                                axis=alt.Axis(format=",.0f")),
                        color=alt.Color("방향:N", scale=alt.Scale(
                            domain=["실적↑","실적↓"],
                            range=["#198754","#dc3545"])),
                        tooltip=["월:N", alt.Tooltip("차이(만):Q", format=",.1f")],
                    )
                    zero_l2 = alt.Chart(pd.DataFrame({"y":[0]})).mark_rule(
                        color="gray", strokeDash=[4,4]).encode(y="y:Q")
                    st.altair_chart(
                        (bar_diff + zero_l2).properties(
                            title="영업손익 차이: 실적 − 시뮬 (만원, 초록=실적우세)",
                            height=200),
                        use_container_width=True)

                # 주요 비용 항목별 갭 히트맵/테이블
                st.markdown("**🔍 주요 비용 항목 갭 (실적 − 시뮬, 만원)**")
                COST_ITEMS = ["매출원가","광고선전비","판매수수료","운반비","지급수수료","수출제비용","판관비계"]
                gap_rows = []
                for m in compare_months:
                    s   = sim_results[m]
                    act = fin_hist[m]["data"]
                    row_d = {"월": m}
                    for itm in COST_ITEMS:
                        sv = s.get(itm, 0)
                        av = act.get(itm, 0)
                        row_d[itm] = f"{(av-sv)/10000:+,.0f}"
                    gap_rows.append(row_d)
                st.dataframe(pd.DataFrame(gap_rows), use_container_width=True, hide_index=True)


    with tab6:
        st.markdown("""
## 🔧 Google Sheets 연동 설정 가이드

### 1단계 — Google Cloud 설정

1. [Google Cloud Console](https://console.cloud.google.com) 접속
2. 새 프로젝트 생성 또는 기존 프로젝트 선택
3. **API 및 서비스 → 라이브러리** 에서 **Google Sheets API** 활성화

### 2단계 — 서비스 계정 생성

1. **API 및 서비스 → 사용자 인증 정보** → 서비스 계정 만들기
2. 키 탭 → **JSON 키 추가** → 다운로드
3. 다운로드한 파일을 `credentials.json` 으로 이름 바꿔 이 폴더에 저장

### 3단계 — 구글 시트 공유

구글 시트를 열고 **공유** → 서비스 계정 이메일을 **편집자** 권한으로 추가

### 4단계 — 앱 실행

```
run.bat 더블클릭
```

### 5단계 — 임직원 공유

| 방법 | 특징 | 권장 |
|------|------|------|
| 사내 PC에서 실행 후 IP 공유 | `http://192.168.x.x:8501` | 내부망 공유 |
| Streamlit Community Cloud | 무료, 공개 URL | 외부 공유 |
| ngrok 터널링 | 임시 URL | 테스트용 |

> **Streamlit Cloud 배포 시** `credentials.json` 대신
> **Secrets 관리** 탭에 JSON 내용을 `[google_credentials]` 형식으로 입력하세요.
        """)


if __name__ == "__main__":
    main()
